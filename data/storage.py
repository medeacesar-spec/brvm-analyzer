"""
Stockage persistant pour les données fondamentales, portefeuille, profil.

La couche de connexion (SQLite local OU Postgres cloud) est abstraite dans
`data.db.get_connection()`. Le code de ce module n'est pas dépendant du SGBD.
"""

import json
import os
import sqlite3
from typing import Optional

import pandas as pd

from config import DB_PATH
from data.db import get_connection, current_user_id, read_sql_df  # noqa: F401 — réexportés


def _maybe_cache_data(ttl: int = 300):
    """Décorateur qui utilise st.cache_data si Streamlit est dispo, sinon no-op.
    Permet de mémoiser un résultat DB pendant `ttl` secondes côté process,
    évitant les aller-retours Supabase lors de la navigation entre pages."""
    try:
        import streamlit as _st
        return _st.cache_data(ttl=ttl, show_spinner=False)
    except Exception:
        def _no_op(fn):
            return fn
        return _no_op


def _resolve_user(user_id: Optional[str]) -> str:
    """Retourne le user_id donné ou celui de l'utilisateur courant ('local' par défaut)."""
    return user_id if user_id else current_user_id()


def init_db():
    """Crée les tables si elles n'existent pas."""
    conn = get_connection()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS fundamentals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            company_name TEXT,
            sector TEXT,
            currency TEXT DEFAULT 'XOF',
            fiscal_year INTEGER,
            price REAL,
            shares REAL,
            float_pct REAL,
            revenue REAL,
            net_income REAL,
            equity REAL,
            total_debt REAL,
            total_assets REAL,
            ebit REAL,
            interest_expense REAL,
            cfo REAL,
            capex REAL,
            dividends_total REAL,
            dps REAL,
            eps REAL,
            per REAL,
            revenue_growth REAL,
            net_income_growth REAL,
            -- Historique N-3 à N
            revenue_n3 REAL,
            revenue_n2 REAL,
            revenue_n1 REAL,
            revenue_n0 REAL,
            net_income_n3 REAL,
            net_income_n2 REAL,
            net_income_n1 REAL,
            net_income_n0 REAL,
            dps_n3 REAL,
            dps_n2 REAL,
            dps_n1 REAL,
            dps_n0 REAL,
            -- Métadonnées
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, fiscal_year)
        );

        CREATE TABLE IF NOT EXISTS portfolio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            company_name TEXT,
            quantity REAL NOT NULL,
            avg_price REAL NOT NULL,
            purchase_date TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS price_cache (
            ticker TEXT NOT NULL,
            date TEXT NOT NULL,
            open REAL,
            high REAL,
            low REAL,
            close REAL,
            volume REAL,
            PRIMARY KEY (ticker, date)
        );

        CREATE TABLE IF NOT EXISTS investor_profile (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            risk_profile TEXT,
            horizon TEXT,
            budget REAL,
            preferred_sectors TEXT,
            preferred_tickers TEXT,
            excluded_tickers TEXT,
            objective TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS indices_cache (
            name TEXT PRIMARY KEY,
            value REAL,
            variation REAL,
            prev_close REAL,
            ytd_variation REAL,
            category TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS market_data (
            ticker TEXT PRIMARY KEY,
            company_name TEXT,
            sector TEXT,
            price REAL,
            variation REAL,
            market_cap REAL,
            beta REAL,
            rsi REAL,
            dps REAL,
            dividend_yield REAL,
            dividend_history TEXT,
            shares REAL,
            float_pct REAL,
            per REAL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS publications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            title TEXT,
            pub_type TEXT,
            period TEXT,
            url TEXT,
            pub_date TEXT,
            is_new INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, title)
        );

        CREATE TABLE IF NOT EXISTS quarterly_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            fiscal_year INTEGER,
            quarter INTEGER,
            revenue REAL,
            net_income REAL,
            ebit REAL,
            source TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, fiscal_year, quarter)
        );

        CREATE TABLE IF NOT EXISTS qualitative_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            category TEXT,
            content TEXT NOT NULL,
            source TEXT,
            note_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Écarts de publication ignorés manuellement (non applicables)
        CREATE TABLE IF NOT EXISTS ignored_gaps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            gap_type TEXT NOT NULL,      -- 'annuel' | 'trimestriel'
            fiscal_year INTEGER,
            reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, gap_type, fiscal_year)
        );

        CREATE TABLE IF NOT EXISTS report_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            title TEXT,
            report_type TEXT,
            fiscal_year INTEGER,
            url TEXT NOT NULL,
            source TEXT DEFAULT 'brvm.org',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, url)
        );

        CREATE TABLE IF NOT EXISTS company_profiles (
            ticker TEXT PRIMARY KEY,
            description TEXT,
            business TEXT,
            president TEXT,
            dg TEXT,
            dga TEXT,
            phone TEXT,
            fax TEXT,
            address TEXT,
            website TEXT,
            major_shareholder TEXT,
            major_shareholder_pct REAL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS company_news (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            title TEXT NOT NULL,
            url TEXT,
            article_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ticker, title)
        );

        CREATE TABLE IF NOT EXISTS signal_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticker TEXT NOT NULL,
            company_name TEXT,
            sector TEXT,
            entry_type TEXT NOT NULL,
            signal_type TEXT,
            signal_name TEXT,
            verdict TEXT,
            first_seen_date TEXT NOT NULL,
            last_seen_date TEXT NOT NULL,
            occurrence_count INTEGER DEFAULT 1,
            price_at_start REAL,
            signal_strength INTEGER,
            signal_details TEXT,
            hybrid_score REAL,
            fundamental_score REAL,
            technical_score REAL,
            stars INTEGER,
            trend TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_signal_history_ticker
            ON signal_history(ticker, last_seen_date DESC);
        CREATE INDEX IF NOT EXISTS idx_signal_history_event
            ON signal_history(ticker, entry_type, signal_name, verdict, last_seen_date DESC);

        CREATE TABLE IF NOT EXISTS portfolio_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Snapshots mensuels des poids calibrés (historique pour suivre l'évolution)
        CREATE TABLE IF NOT EXISTS calibration_reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            review_date TEXT NOT NULL,
            days_available INTEGER,
            n_signals INTEGER,
            n_recos INTEGER,
            calibrated_signals INTEGER,
            calibrated_recos INTEGER,
            payload TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_cal_reviews_date
            ON calibration_reviews(review_date DESC);

        -- ── SNAPSHOTS QUOTIDIENS (pré-calculés par le cron GitHub Actions) ──
        -- Évitent les N+1 et le compute live sur les pages Signaux, Performance,
        -- Historique. Les pages deviennent des SELECT triviaux (<1 s).

        CREATE TABLE IF NOT EXISTS scoring_snapshot (
            ticker TEXT PRIMARY KEY,
            company_name TEXT,
            sector TEXT,
            price REAL,
            hybrid_score REAL,
            fundamental_score REAL,
            technical_score REAL,
            verdict TEXT,
            stars INTEGER,
            trend TEXT,
            nb_signals INTEGER,
            signals_json TEXT,
            consolidated_json TEXT,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Journal quotidien append-only des verdicts/scores. Source de vérité
        -- pour : trajectoires, backtests, hit rate, score evolution, cohorts.
        -- ~12k lignes/an pour 48 tickers, négligeable. PK composite (ticker,date)
        -- garantit l'idempotence si build_daily_snapshot tourne 2x dans la journée.
        CREATE TABLE IF NOT EXISTS verdict_daily (
            ticker TEXT NOT NULL,
            date TEXT NOT NULL,
            verdict TEXT,
            stars INTEGER,
            hybrid_score REAL,
            fundamental_score REAL,
            technical_score REAL,
            price REAL,
            trend TEXT,
            nb_signals INTEGER,
            sector TEXT,
            company_name TEXT,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (ticker, date)
        );

        CREATE INDEX IF NOT EXISTS idx_verdict_daily_date ON verdict_daily(date);
        CREATE INDEX IF NOT EXISTS idx_verdict_daily_verdict ON verdict_daily(verdict);

        CREATE TABLE IF NOT EXISTS ticker_performance_snapshot (
            ticker TEXT PRIMARY KEY,
            company_name TEXT,
            sector TEXT,
            last_price REAL,
            last_date TEXT,
            perf_1m REAL,
            perf_3m REAL,
            perf_6m REAL,
            perf_1a REAL,
            perf_2a REAL,
            perf_3a REAL,
            perf_max REAL,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS signal_performance_snapshot (
            event_id INTEGER PRIMARY KEY,
            current_price REAL,
            perf_1m REAL,
            perf_3m REAL,
            perf_6m REAL,
            perf_1a REAL,
            perf_since_start REAL,
            duration_days INTEGER,
            computed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Métadonnées du dernier build (pour afficher la fraîcheur dans l'UI)
        CREATE TABLE IF NOT EXISTS snapshot_meta (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    # Ensure 'ignored' column exists on publications (for old DBs)
    cols = _table_columns(conn, "publications")
    if "ignored" not in cols:
        _safe_alter(conn, "ALTER TABLE publications ADD COLUMN ignored INTEGER DEFAULT 0")
    if "source" not in cols:
        _safe_alter(conn, "ALTER TABLE publications ADD COLUMN source TEXT")
    if "fiscal_year" not in cols:
        _safe_alter(conn, "ALTER TABLE publications ADD COLUMN fiscal_year INTEGER")

    # Réconciliation de schéma : colonnes ajoutées après la migration initiale.
    # Évite les "UndefinedColumn" sur Postgres où CREATE TABLE IF NOT EXISTS
    # est no-op si la table existait déjà avec un schéma partiel.
    _fund_cols = _table_columns(conn, "fundamentals")
    for col, ctype in [
        ("total_assets", "REAL"), ("eps", "REAL"), ("per", "REAL"),
        ("revenue_growth", "REAL"), ("net_income_growth", "REAL"),
        ("shares", "REAL"), ("float_pct", "REAL"),
    ]:
        if col not in _fund_cols:
            _safe_alter(conn, f"ALTER TABLE fundamentals ADD COLUMN {col} REAL")
    _md_cols = _table_columns(conn, "market_data")
    for col, ctype in [
        ("shares", "REAL"), ("float_pct", "REAL"),
        ("dividend_yield", "REAL"), ("per", "REAL"),
    ]:
        if col not in _md_cols:
            _safe_alter(conn, f"ALTER TABLE market_data ADD COLUMN {col} REAL")

    # ─────────────────────────────────────────────────────────────
    # Migration multi-utilisateurs : ajout de user_id sur les tables
    # propres à chaque utilisateur (portfolio, cash, profil, notes).
    # Idempotent : se contente de ne rien faire si déjà appliqué.
    # ─────────────────────────────────────────────────────────────
    _migrate_user_scoped_tables(conn)

    conn.commit()
    conn.close()


def _table_columns(conn, table: str) -> list:
    """Retourne la liste des noms de colonnes d'une table — compatible SQLite & Postgres."""
    try:
        from data.db import is_postgres
        if is_postgres():
            rows = conn.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name = ? ORDER BY ordinal_position",
                (table,),
            ).fetchall()
            # dict_row → col name key varies
            out = []
            for r in rows:
                if isinstance(r, dict):
                    out.append(r.get("column_name"))
                else:
                    out.append(r[0])
            return out
    except Exception:
        pass
    # SQLite fallback
    return [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]


def _safe_alter(conn, sql: str):
    """Exécute un ALTER TABLE en ignorant l'erreur si la colonne existe déjà."""
    try:
        conn.execute(sql)
    except Exception:
        pass


def _migrate_user_scoped_tables(conn):
    """Ajoute `user_id` aux tables par-utilisateur. Sûr à relancer."""
    def cols_of(table):
        return _table_columns(conn, table)

    # 1) portfolio : simple ALTER TABLE ADD COLUMN
    if "user_id" not in cols_of("portfolio"):
        conn.execute(
            "ALTER TABLE portfolio ADD COLUMN user_id TEXT NOT NULL DEFAULT 'local'"
        )

    # 2) qualitative_notes : idem
    if "user_id" not in cols_of("qualitative_notes"):
        conn.execute(
            "ALTER TABLE qualitative_notes ADD COLUMN user_id TEXT NOT NULL DEFAULT 'local'"
        )

    # 3) portfolio_settings : ancienne PK = (key). Nouvelle PK = (user_id, key).
    #    On doit recréer la table.
    if "user_id" not in cols_of("portfolio_settings"):
        conn.executescript("""
            CREATE TABLE portfolio_settings_new (
                user_id TEXT NOT NULL DEFAULT 'local',
                key TEXT NOT NULL,
                value TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, key)
            );
            INSERT INTO portfolio_settings_new (user_id, key, value, updated_at)
                SELECT 'local', key, value, updated_at FROM portfolio_settings;
            DROP TABLE portfolio_settings;
            ALTER TABLE portfolio_settings_new RENAME TO portfolio_settings;
        """)

    # 4) investor_profile : ancienne PK = (id=1 singleton). Nouvelle PK = user_id.
    if "user_id" not in cols_of("investor_profile"):
        conn.executescript("""
            CREATE TABLE investor_profile_new (
                user_id TEXT PRIMARY KEY,
                risk_profile TEXT,
                horizon TEXT,
                budget REAL,
                preferred_sectors TEXT,
                preferred_tickers TEXT,
                excluded_tickers TEXT,
                objective TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            INSERT INTO investor_profile_new
                (user_id, risk_profile, horizon, budget,
                 preferred_sectors, preferred_tickers, excluded_tickers,
                 objective, updated_at)
                SELECT 'local', risk_profile, horizon, budget,
                       preferred_sectors, preferred_tickers, excluded_tickers,
                       objective, updated_at FROM investor_profile;
            DROP TABLE investor_profile;
            ALTER TABLE investor_profile_new RENAME TO investor_profile;
        """)


def get_analyzable_tickers() -> list:
    """
    Retourne les tickers qui ont assez de donnees pour etre analyses.
    Un titre est analysable s'il a des fondamentaux (Excel import)
    OU des donnees de marche avec prix + DPS.

    Noms et secteurs : fallback sur brvm_tickers.json si la DB contient
    company_name/sector à NULL (cas de market_data tronqué).
    """
    from config import load_tickers as _load_cfg
    conn = get_connection()
    rows = conn.execute("""
        SELECT DISTINCT ticker, company_name, sector, 'fundamentals' as source
        FROM fundamentals
        UNION
        SELECT DISTINCT ticker, company_name, sector, 'market' as source
        FROM market_data
        WHERE price IS NOT NULL AND price > 0
        ORDER BY ticker
    """).fetchall()
    conn.close()

    # Fallback config : {ticker: {name, sector}} depuis brvm_tickers.json
    cfg = {t["ticker"]: t for t in _load_cfg()}

    # Deduplicate by ticker, prefer fundamentals source
    seen = {}
    for r in rows:
        t = r["ticker"]
        cfg_t = cfg.get(t, {})
        db_name = r["company_name"]
        db_sector = r["sector"]
        # Remplace None/"" par le nom/secteur du config
        name = db_name if (db_name and str(db_name).strip() and str(db_name).lower() != "none") else cfg_t.get("name", t)
        sector = db_sector if (db_sector and str(db_sector).strip()) else cfg_t.get("sector", "")
        if t not in seen or r["source"] == "fundamentals":
            seen[t] = {"ticker": t, "name": name, "sector": sector,
                       "has_fundamentals": r["source"] == "fundamentals"}
    return list(seen.values())


# --- Fundamentals CRUD ---

def _sanitize_fundamentals(data: dict) -> dict:
    """Nettoie les valeurs aberrantes avant insertion.
    Met à NULL toute valeur clairement incohérente pour éviter de polluer la base.

    Seuils calibrés pour la BRVM (le plus gros emetteur Sonatel ~1.6T FCFA
    revenue, ~330B net income, banques peuvent avoir des bilans plus gros).
    """
    data = dict(data)  # copy
    rev = data.get("revenue")
    ni = data.get("net_income")
    eq = data.get("equity")
    debt = data.get("total_debt")

    reasons = []

    def _clear(*keys, why):
        for k in keys:
            data[k] = None
        reasons.append(why)

    # 1. Marge nette absurde. Les banques ont typiquement 40-75% de marge
    # (PNB - charges d'exploitation / PNB), les industriels 5-25%. On adapte
    # le seuil selon le secteur si connu (fallback 80% pour être tolérant).
    sector_str = (data.get("sector") or "").lower()
    is_bank_sec = "banque" in sector_str or "bank" in sector_str
    ticker_str = (data.get("ticker") or "")
    # Heuristique tickers banques BRVM (suffixes .ci/.bj/.tg/.sn courants)
    # Comparaison sur préfixe uppercase pour être insensible à la casse du suffix.
    _BANK_PREFIXES = {
        "BICC", "BICB", "BOAB", "BOABF", "BOAC", "BOAM", "BOAN", "BOAS",
        "CBIBF", "ECOC", "ETIT", "NSBC", "ORGT", "SAFC", "SGBC", "SIBC",
    }
    ticker_prefix = ticker_str.split(".")[0].upper() if ticker_str else ""
    is_bank = is_bank_sec or ticker_prefix in _BANK_PREFIXES

    if rev and ni and rev != 0:
        margin = ni / rev
        # Banques : jusqu'à 80% possible (BICB à 68%, SIBC à 51%).
        # Industriels : max 60%.
        upper_margin = 0.80 if is_bank else 0.60
        if margin > upper_margin or margin < -1.0:
            _clear("revenue", "net_income",
                   why=f"Marge absurde {margin*100:.1f}% "
                        f"(rev={rev:,.0f}, ni={ni:,.0f}, banque={is_bank})")

    # 2. Revenue < 1 million pour un titre coté → probablement tronqué
    if rev is not None and 0 < abs(rev) < 1e6:
        _clear("revenue", why=f"Revenue trop faible : {rev:,.0f}")

    # 3. Equity < 1 million → probablement tronqué
    if eq is not None and 0 < abs(eq) < 1e6:
        _clear("equity", why=f"Equity trop faible : {eq:,.0f}")

    # 4. ROE > 200% → incohérent
    if ni and eq and eq > 0:
        roe = ni / eq
        if abs(roe) > 2.0:
            _clear("equity", why=f"ROE absurde {roe*100:.1f}%")

    # 5. Seuils superieurs par champ — le parser PDF se trompe souvent d'un
    # ordre de grandeur (loupe une virgule, lit une ligne sur deux, etc.)
    UPPER_LIMITS = {
        "revenue": 5e12,           # 5T FCFA (Sonatel ~1.6T)
        "net_income": 1e12,        # 1T FCFA (Sonatel ~330B)
        "ebit": 1e12,
        "equity": 5e12,
        "total_debt": 5e12,
        "cfo": 1e12,
        "capex": 5e11,
        "interest_expense": 5e11,
        "total_assets": 50e12,     # banques ont gros bilan, mais 50T = limite generouse
        "dividends_total": 1e12,
    }
    for key, limit in UPPER_LIMITS.items():
        v = data.get(key)
        if v is not None and abs(v) > limit:
            _clear(key, why=f"{key} > {limit:.0e} FCFA : {v:,.0f}")

    # 6. D/E > 20x hors banques → suspect
    sector = (data.get("sector") or "").lower()
    is_bank = "banque" in sector or "bank" in sector
    if debt and eq and eq != 0 and not is_bank:
        de = debt / eq
        if abs(de) > 20:
            _clear("total_debt", why=f"D/E absurde {de:.1f}x (hors banque)")

    if reasons:
        print(f"[sanitize] {data.get('ticker')} {data.get('fiscal_year')}: "
              + " | ".join(reasons))

    return data


def _net_income_coherent(data: dict) -> dict:
    """Refuse un resultat net entrant qui contredit le BNPA deja en base.

    L'extraction PDF a plusieurs fois ecrase de bons chiffres : Orange CI 2024
    est passe de 158,2 Mds a 112,8 Mds, PALM CI d'un benefice de 15,9 Mds a une
    perte de 619 M. A chaque fois, le BNPA de la fiche societe et le nombre de
    titres — inchanges — disaient le contraire.

    Regle : si la base porte un BNPA et un nombre de titres coherents entre eux,
    et que le resultat net entrant s'en ecarte de plus de 15 %, on ne l'ecrit
    pas. Il n'est pas remplace par un calcul : il est simplement ignore, et
    l'ecart journalise. Un vrai correctif passera par la saisie manuelle ou par
    un rescraping de la fiche.
    """
    ni = data.get("net_income")
    ticker, annee = data.get("ticker"), data.get("fiscal_year")
    if ni is None or not ticker or not annee:
        return data

    try:
        conn = get_connection()
        row = conn.execute(
            "SELECT eps, shares, net_income FROM fundamentals "
            "WHERE ticker = ? AND fiscal_year = ?",
            (ticker, annee),
        ).fetchone()
        conn.close()
    except Exception:
        return data

    if not row:
        return data
    existant = dict(row)
    eps, shares = existant.get("eps"), existant.get("shares")
    if not eps or not shares:
        return data

    attendu = eps * shares
    if attendu and abs(ni - attendu) / abs(attendu) > 0.15:
        print(f"[coherence] {ticker} {annee} : resultat net entrant "
              f"{ni:,.0f} contredit BNPA x titres ({attendu:,.0f}) — ignore",
              flush=True)
        data = dict(data)
        data["net_income"] = None
    return data


# Champs que l'extracteur PDF sait produire et que `save_fundamentals`
# accepte. Centralises ici volontairement : chaque appelant recopiait
# jusqu'ici sa propre liste en dur, si bien que les colonnes ajoutees
# ensuite — ebitda, cout du risque, depots, credits, frais generaux — etaient
# bien extraites mais jamais ecrites. Les grilles sectorielles restaient donc
# vides alors que l'extraction fonctionnait.
CHAMPS_EXTRACTION = (
    "revenue", "net_income", "equity", "total_debt", "ebit",
    "interest_expense", "cfo", "capex", "dividends_total", "total_assets",
    "shares", "ordinary_income", "hao_income", "cost_of_risk", "ebitda",
    "operating_expenses", "gross_operating_income", "pretax_income",
    "deposits", "loans",
)


def champs_extraits(result: dict, selecteur=None) -> dict:
    """Reprend d'un resultat d'extraction tous les champs renseignes.

    `selecteur` permet aux appelants qui arbitrent entre plusieurs sources
    (fonction `_best`) de conserver leur arbitrage.
    """
    prendre = selecteur or (lambda champ: result.get(champ))
    sortie = {}
    for champ in CHAMPS_EXTRACTION:
        try:
            valeur = prendre(champ)
        except Exception:
            valeur = None
        if valeur is not None:
            sortie[champ] = valeur
    return sortie


_SECTEURS_CACHE = {}


def _secteur_connu(ticker: str):
    """Retrouve le secteur d'un ticker : market_data d'abord, referentiel ensuite.

    Sans cela, chaque ligne creee par une extraction arrivait sans secteur —
    et l'affichage des grilles sectorielles, qui en depend, restait muet. Le
    rattrapage du 05/09 a d'ailleurs recree le trou aussitot apres l'avoir
    comble, en insérant des exercices neufs sans secteur.
    """
    if not ticker:
        return None
    if not _SECTEURS_CACHE:
        try:
            conn = get_connection()
            for ligne in conn.execute(
                "SELECT DISTINCT ticker, sector FROM market_data "
                "WHERE sector IS NOT NULL AND sector <> ''"
            ).fetchall():
                d = dict(ligne)
                _SECTEURS_CACHE[d["ticker"]] = (d["sector"] or "").strip()
            conn.close()
        except Exception:
            pass
        try:
            import json as _json
            chemin = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  "brvm_tickers.json")
            with open(chemin, encoding="utf-8") as fh:
                brut = _json.load(fh)
            paires = (brut.items() if isinstance(brut, dict)
                      else [(x.get("ticker"), x) for x in brut])
            for t, v in paires:
                if t and not _SECTEURS_CACHE.get(t):
                    _SECTEURS_CACHE[t] = (v.get("sector") or "").strip()
        except Exception:
            pass
    return _SECTEURS_CACHE.get(ticker) or None


def save_fundamentals(data: dict) -> int:
    """Insère ou met à jour les données fondamentales d'un titre.
    Les valeurs clairement aberrantes sont mises à NULL avant insertion.
    En UPDATE, les champs NULL (non fournis) n'écrasent PAS les valeurs existantes :
    on complète seulement ce qui manque (COALESCE)."""
    data = _sanitize_fundamentals(data)
    data = _net_income_coherent(data)
    if not (data.get("sector") or "").strip():
        _sect = _secteur_connu(data.get("ticker"))
        if _sect:
            data["sector"] = _sect
    conn = get_connection()
    cols = [
        "ticker", "company_name", "sector", "currency", "fiscal_year",
        "price", "shares", "float_pct", "revenue", "net_income", "equity",
        "total_debt", "ebit", "interest_expense", "cfo", "capex",
        "dividends_total", "dps", "total_assets", "eps", "per",
        "ordinary_income", "hao_income", "cost_of_risk",
        "ebitda",
        "operating_expenses", "gross_operating_income", "pretax_income",
        "deposits", "loans",
        "revenue_n3", "revenue_n2", "revenue_n1", "revenue_n0",
        "net_income_n3", "net_income_n2", "net_income_n1", "net_income_n0",
        "dps_n3", "dps_n2", "dps_n1", "dps_n0",
    ]
    values = [data.get(c) for c in cols]

    placeholders = ", ".join(["?"] * len(cols))
    col_names = ", ".join(cols)
    # On ne remplace une valeur existante que si la nouvelle n'est pas NULL.
    # Pour les fields 'metadata' (company_name, sector, currency), on permet le remplacement.
    _update_always = {"company_name", "sector", "currency"}
    _skip = {"ticker", "fiscal_year"}
    update_parts = []
    for c in cols:
        if c in _skip:
            continue
        if c in _update_always:
            update_parts.append(f"{c}=excluded.{c}")
        else:
            # COALESCE(excluded.col, fundamentals.col) → prend la nouvelle
            # valeur si non-NULL, sinon garde l'ancienne. La table doit etre
            # qualifiee : Postgres rejette `price` ambigu entre la nouvelle
            # ligne et la cible du SET (SQLite etait tolerant).
            update_parts.append(
                f"{c}=COALESCE(excluded.{c}, fundamentals.{c})"
            )
    update_clause = ", ".join(update_parts)

    cursor = conn.execute(
        f"""INSERT INTO fundamentals ({col_names})
            VALUES ({placeholders})
            ON CONFLICT(ticker, fiscal_year) DO UPDATE SET
            {update_clause}, updated_at=CURRENT_TIMESTAMP""",
        values,
    )
    conn.commit()
    row_id = cursor.lastrowid

    # Auto-clearance : marque les publications annuelles/EF correspondantes
    # comme traitées si les données saisies sont substantielles.
    ticker = data.get("ticker")
    fy = data.get("fiscal_year")
    if ticker and fy and data.get("revenue") is not None:
        try:
            _mark_pubs_processed(conn, ticker, fy,
                                  ("annuel", "etats_financiers"))
            conn.commit()
        except Exception:
            pass  # non bloquant

    conn.close()
    return row_id


TYPES_ANNUELS = ("rapport_annuel", "etats_financiers")


def exercices_annuels(tickers=None) -> dict:
    """Exercices pour lesquels un document ANNUEL est reference.

    C'est le seul signal qui ne se devine pas. Le cycle de publication UEMOA
    veut qu'un exercice N paraisse au printemps N+1 ; pendant l'annee N il
    n'existe que des trimestriels et des semestriels. Un exercice sans
    document annuel n'est donc pas un exercice clos, quoi qu'en dise la table
    des fondamentaux.

    Une seule requete, pour que la fiche d'un titre et la comparaison
    sectorielle s'accordent toujours sur la meme annee — c'est precisement
    leur desaccord qui faisait apparaitre la SIB sur 2024 dans le tableau des
    pairs alors que son exercice 2025 est publie.
    """
    conn = get_connection()
    try:
        marques = ",".join("?" * len(TYPES_ANNUELS))
        if tickers:
            liste = list(tickers)
            place = ",".join("?" * len(liste))
            lignes = conn.execute(
                f"""SELECT DISTINCT ticker, fiscal_year FROM report_links
                    WHERE report_type IN ({marques}) AND ticker IN ({place})""",
                (*TYPES_ANNUELS, *liste),
            ).fetchall()
        else:
            lignes = conn.execute(
                f"""SELECT DISTINCT ticker, fiscal_year FROM report_links
                    WHERE report_type IN ({marques})""",
                TYPES_ANNUELS,
            ).fetchall()
    except Exception:
        conn.close()
        return {}
    conn.close()

    sortie = {}
    for ligne in lignes:
        d = dict(ligne)
        annee = d.get("fiscal_year")
        if d.get("ticker") and annee:
            sortie.setdefault(d["ticker"], set()).add(int(annee))
    return sortie


def get_fundamentals(ticker: str, fiscal_year: Optional[int] = None) -> Optional[dict]:
    """Récupère les données fondamentales d'un titre.
    Sans fiscal_year, prend la dernière année ayant revenue + net_income.
    Si equity, total_debt, cfo ou capex manquent sur cette année, on les complète
    avec la valeur de l'année la plus récente qui les a (proxy utile sinon l'analyse
    ratio basée bilan est impossible alors qu'on a l'income statement récent)."""
    conn = get_connection()
    if fiscal_year:
        row = conn.execute(
            "SELECT * FROM fundamentals WHERE ticker=? AND fiscal_year=?",
            (ticker, fiscal_year),
        ).fetchone()
    else:
        # Priorité : dernière année avec revenue + net_income (income statement
        # complet) — MAIS en écartant les exercices manifestement partiels.
        #
        # Aucun émetteur ne publie de rapport annuel 2026 en septembre 2026 :
        # les seules publications de l'année sont trimestrielles ou
        # semestrielles. Leur chiffre d'affaires et leur résultat net couvrent
        # donc un trimestre, et la fiche les présentait comme un exercice.
        # Orange CI affichait ainsi un PER de 65,5 — son bénéfice trimestriel
        # rapporté au cours, soit environ quatre fois le PER réel.
        #
        # On repère ces exercices par leur ordre de grandeur : un chiffre
        # d'affaires inférieur à la moitié de la médiane du titre n'est pas
        # une année, c'est une fraction d'année.
        candidats = [dict(l) for l in conn.execute(
            """SELECT * FROM fundamentals WHERE ticker=?
               AND revenue IS NOT NULL AND revenue != 0
               AND net_income IS NOT NULL AND net_income != 0
               ORDER BY fiscal_year DESC""",
            (ticker,),
        ).fetchall()]

        row = None
        if candidats:
            # Premier critère, le seul qui ne se devine pas : l'exercice
            # dispose-t-il d'un document ANNUEL ? Le cycle de publication
            # UEMOA veut qu'un exercice N soit publié au printemps N+1 ;
            # pendant l'année N, il n'existe que des trimestriels et des
            # semestriels. Un exercice sans document annuel n'est donc pas un
            # exercice clos, quoi qu'en dise la table.
            annuels = exercices_annuels([ticker]).get(ticker, set())

            for candidat in candidats:
                if candidat.get("fiscal_year") in annuels:
                    row = candidat
                    break

            # Second critère, de repli quand aucun document n'est référencé :
            # l'ordre de grandeur. Un chiffre d'affaires inférieur à la moitié
            # de la médiane du titre n'est pas une année, c'est une fraction
            # d'année.
            if row is None:
                montants = sorted(abs(c["revenue"]) for c in candidats
                                  if c.get("revenue"))
                repere = (montants[len(montants) // 2] if montants else None)
                for candidat in candidats:
                    if (repere and candidat.get("revenue")
                            and abs(candidat["revenue"]) < 0.5 * repere):
                        continue
                    row = candidat
                    break

            # Rien de concluant : on garde le plus récent plutôt que rien,
            # mais on le signale pour que l'écran puisse le dire.
            if row is None:
                row = dict(candidats[0])
                row["_exercice_incertain"] = True
        if not row:
            row = conn.execute(
                """SELECT * FROM fundamentals WHERE ticker=?
                   AND (revenue IS NOT NULL AND revenue != 0
                        OR net_income IS NOT NULL AND net_income != 0)
                   ORDER BY fiscal_year DESC LIMIT 1""",
                (ticker,),
            ).fetchone()
        if not row:
            row = conn.execute(
                "SELECT * FROM fundamentals WHERE ticker=? ORDER BY fiscal_year DESC LIMIT 1",
                (ticker,),
            ).fetchone()

        # Complète les champs bilan/cashflow manquants depuis l'année la plus récente qui les a
        if row:
            row_dict = dict(row)
            ref_year = row_dict.get("fiscal_year")
            _cols_avail = set(_table_columns(conn, "fundamentals"))
            for fld in ("equity", "total_debt", "total_assets", "cfo", "capex",
                        "ebit", "interest_expense", "dividends_total"):
                if fld not in _cols_avail:
                    continue  # colonne absente de cette DB → skip
                if row_dict.get(fld) is None:
                    try:
                        proxy = conn.execute(
                            f"""SELECT {fld}, fiscal_year FROM fundamentals
                               WHERE ticker=? AND {fld} IS NOT NULL AND {fld} != 0
                               AND fiscal_year <= ?
                               ORDER BY fiscal_year DESC LIMIT 1""",
                            (ticker, ref_year),
                        ).fetchone()
                    except Exception:
                        proxy = None
                    if proxy:
                        row_dict[fld] = proxy[0]
                        row_dict[f"_{fld}_from_year"] = proxy[1]

            # Recalcule dynamiquement l'historique N-3..N0 à partir des 4 années
            # les plus récentes (≤ ref_year) disponibles dans la DB.
            hist_rows = conn.execute(
                """SELECT fiscal_year, revenue, net_income, dps FROM fundamentals
                   WHERE ticker = ?
                   AND (revenue IS NOT NULL AND revenue != 0
                        OR net_income IS NOT NULL AND net_income != 0)
                   AND fiscal_year <= ?
                   ORDER BY fiscal_year DESC LIMIT 4""",
                (ticker, ref_year),
            ).fetchall()
            for idx, hr in enumerate(hist_rows):
                suf = ["n0", "n1", "n2", "n3"][idx]
                if hr["revenue"]:
                    row_dict[f"revenue_{suf}"] = hr["revenue"]
                if hr["net_income"]:
                    row_dict[f"net_income_{suf}"] = hr["net_income"]
                if hr["dps"]:
                    row_dict[f"dps_{suf}"] = hr["dps"]
            row = row_dict
    conn.close()
    if not row:
        return None

    result = dict(row)

    # Enrichissement depuis market_data.
    #
    # Le cours porte par `fundamentals` est celui du jour ou la fiche a ete
    # scrapee, pas un cours d'exercice : il vieillit sans jamais etre rafraichi.
    # Tant qu'on ne le consultait qu'a defaut, il masquait le cours reel — 36
    # tickers sur 48 s'en trouvaient decales de plus de 10 %, et jusqu'a 55 %
    # pour ETIT (32 F en base contre 71 F au marche). Comme PER, P/B et
    # rendement se calculent tous sur le cours du jour, market_data fait
    # desormais foi des qu'il a une valeur.
    conn2 = get_connection()
    md = conn2.execute(
        "SELECT price, market_cap, dps, shares, dividend_yield FROM market_data WHERE ticker=?",
        (ticker,),
    ).fetchone()
    conn2.close()
    if md:
        md_dict = dict(md)
        if md_dict.get("price"):
            result["price"] = md_dict["price"]
        # La capitalisation suit le cours : la reprendre aussi, sinon elle
        # resterait coherente avec l'ancien cours et plus avec le nouveau.
        if md_dict.get("market_cap"):
            result["market_cap"] = md_dict["market_cap"]
        elif md_dict.get("price") and result.get("shares"):
            result["market_cap"] = md_dict["price"] * result["shares"]
        if not result.get("shares") and md_dict.get("shares"):
            result["shares"] = md_dict["shares"]
        # Use latest DPS from market_data if fundamentals DPS is missing
        if not result.get("dps") and md_dict.get("dps"):
            result["dps"] = md_dict["dps"]

    return result


def _best_year_subquery() -> str:
    """Sous-requête SQL pour trouver la meilleure année par ticker.
    Priorité : dernière année avec revenue+net_income (income statement complet).
    Equity n'est PLUS un critère de sélection : si manquant, on le récupère d'une
    année antérieure via get_all_stocks_for_analysis (fallback Python).
    """
    return """
        SELECT ticker,
            COALESCE(
                (SELECT MAX(fiscal_year) FROM fundamentals f2
                 WHERE f2.ticker = f1.ticker
                   AND f2.revenue IS NOT NULL AND f2.revenue != 0
                   AND f2.net_income IS NOT NULL AND f2.net_income != 0),
                (SELECT MAX(fiscal_year) FROM fundamentals f2
                 WHERE f2.ticker = f1.ticker
                   AND (f2.revenue IS NOT NULL AND f2.revenue != 0
                        OR f2.net_income IS NOT NULL AND f2.net_income != 0)),
                MAX(fiscal_year)
            ) as max_year
        FROM fundamentals f1
        GROUP BY ticker
    """


def get_all_fundamentals() -> pd.DataFrame:
    """Récupère toutes les données fondamentales (meilleure année par titre)."""
    conn = get_connection()
    df = read_sql_df(f"""SELECT f.* FROM fundamentals f
           INNER JOIN ({_best_year_subquery()}) latest
           ON f.ticker = latest.ticker AND f.fiscal_year = latest.max_year
           ORDER BY f.ticker""")
    conn.close()
    return df


def list_tickers_with_fundamentals() -> list:
    """Liste les tickers qui ont des données fondamentales."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT ticker, company_name FROM fundamentals ORDER BY ticker"
    ).fetchall()
    conn.close()
    return [{"ticker": r["ticker"], "name": r["company_name"]} for r in rows]


# --- Excel Import ---

def import_from_excel(filepath: str) -> dict:
    """
    Importe les données depuis un fichier Excel au format 'Analyse Hybride'.
    Lit la feuille 'Inputs' et retourne un dict compatible avec save_fundamentals.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Inputs"]

    def cell_val(row, col=2):
        return ws.cell(row=row, column=col).value

    data = {
        "company_name": cell_val(4),
        "ticker": _resolve_ticker(cell_val(5)),
        "sector": cell_val(6),
        "currency": cell_val(7) or "XOF",
        "fiscal_year": cell_val(8),
        "price": cell_val(9),
        "shares": cell_val(10),
        "revenue": cell_val(13),
        "net_income": cell_val(14),
        "equity": cell_val(15),
        "total_debt": cell_val(16),
        "ebit": cell_val(17),
        "interest_expense": cell_val(18),
        "cfo": cell_val(19),
        "capex": cell_val(20),
        "dividends_total": cell_val(21),
        "dps": cell_val(22),
    }

    # Parse float percentage from column C row 10
    float_text = ws.cell(row=10, column=3).value
    if float_text and isinstance(float_text, str) and "%" in float_text:
        try:
            data["float_pct"] = float(float_text.replace("%", "").replace(",", ".").strip())
        except ValueError:
            data["float_pct"] = None
    else:
        data["float_pct"] = None

    # Historical data (columns C=N-3, D=N-2, E=N-1, F=N)
    hist_cols = {3: "n3", 4: "n2", 5: "n1", 6: "n0"}
    for col_idx, suffix in hist_cols.items():
        data[f"revenue_{suffix}"] = ws.cell(row=26, column=col_idx).value
        data[f"net_income_{suffix}"] = ws.cell(row=27, column=col_idx).value
        data[f"dps_{suffix}"] = ws.cell(row=28, column=col_idx).value

    wb.close()
    return data


def _resolve_ticker(mnemo: str) -> str:
    """Convertit un mnémonique (ECOC) en ticker complet (ECOC.ci)."""
    if not mnemo:
        return ""
    mnemo = mnemo.strip()
    if "." in mnemo:
        return mnemo

    # Map common mnemonics to full tickers
    from config import load_tickers
    tickers = load_tickers()
    for t in tickers:
        if t["ticker"].split(".")[0].upper() == mnemo.upper():
            return t["ticker"]
    return mnemo


# --- Price Cache ---

def cache_prices(ticker: str, df: pd.DataFrame):
    """Stocke les prix historiques en cache SQLite."""
    if df.empty:
        return
    conn = get_connection()
    for _, row in df.iterrows():
        conn.execute(
            """INSERT INTO price_cache (ticker, date, open, high, low, close, volume)
               VALUES (?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(ticker, date) DO UPDATE SET
                 open=excluded.open, high=excluded.high, low=excluded.low,
                 close=excluded.close, volume=excluded.volume""",
            (
                ticker,
                row["date"].strftime("%Y-%m-%d") if hasattr(row["date"], "strftime") else str(row["date"]),
                row.get("open"),
                row.get("high"),
                row.get("low"),
                row.get("close"),
                row.get("volume"),
            ),
        )
    conn.commit()
    conn.close()


def get_cached_prices(ticker: str) -> pd.DataFrame:
    """Récupère les prix en cache pour un ticker."""
    conn = get_connection()
    df = read_sql_df("SELECT date, open, high, low, close, volume FROM price_cache WHERE ticker=? ORDER BY date", params=(ticker,),
        parse_dates=["date"],
    )
    conn.close()
    return df


@_maybe_cache_data(ttl=300)
def get_all_cached_prices() -> dict:
    """Retourne {ticker: DataFrame} pour TOUS les tickers en une seule requête.
    Utile pour les pages qui bouclent sur tous les titres (Signaux, Comparateur)
    → évite ~48 round-trips Supabase en réduisant à 1."""
    conn = get_connection()
    try:
        df = read_sql_df(
            "SELECT ticker, date, open, high, low, close, volume "
            "FROM price_cache ORDER BY ticker, date",
            parse_dates=["date"],
        )
    finally:
        conn.close()
    if df.empty:
        return {}
    return {tkr: grp.drop(columns=["ticker"]).reset_index(drop=True)
            for tkr, grp in df.groupby("ticker")}


# --- Portfolio (user-scoped) ---

def save_position(ticker: str, company_name: str, quantity: float, avg_price: float,
                  purchase_date: str = None, notes: str = None,
                  user_id: Optional[str] = None, fees: float = 0) -> int:
    """Enregistre une position. `fees` = frais totaux liés à cet achat
    (commission de courtage + frais BRVM + TVA). Le coût de revient réel
    d'une ligne = quantity × avg_price + fees."""
    uid = _resolve_user(user_id)
    conn = get_connection()
    cursor = conn.execute(
        """INSERT INTO portfolio
           (user_id, ticker, company_name, quantity, avg_price, purchase_date,
            notes, fees)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (uid, ticker, company_name, quantity, avg_price, purchase_date, notes,
         float(fees or 0)),
    )
    conn.commit()
    row_id = cursor.lastrowid
    conn.close()
    return row_id


def get_portfolio(user_id: Optional[str] = None) -> pd.DataFrame:
    uid = _resolve_user(user_id)
    conn = get_connection()
    df = read_sql_df("SELECT * FROM portfolio WHERE user_id = ? ORDER BY ticker", params=(uid,),
    )
    conn.close()
    return df


def delete_position(position_id: int, user_id: Optional[str] = None):
    """Supprime une position. Si user_id fourni, vérifie que la position
    appartient bien à cet utilisateur (sécurité)."""
    uid = _resolve_user(user_id)
    conn = get_connection()
    conn.execute(
        "DELETE FROM portfolio WHERE id = ? AND user_id = ?",
        (position_id, uid),
    )
    conn.commit()
    conn.close()


def update_position(position_id: int, quantity: float, avg_price: float,
                     user_id: Optional[str] = None,
                     fees: Optional[float] = None) -> bool:
    """Met à jour quantité, PRU et (optionnellement) frais d'une position.
    Ne change ni le ticker ni la date d'achat. Si `fees` est None, la valeur
    existante est préservée. Renvoie True si succès."""
    uid = _resolve_user(user_id)
    conn = get_connection()
    try:
        if fees is None:
            conn.execute(
                "UPDATE portfolio SET quantity = ?, avg_price = ? "
                "WHERE id = ? AND user_id = ?",
                (float(quantity), float(avg_price), int(position_id), uid),
            )
        else:
            conn.execute(
                "UPDATE portfolio SET quantity = ?, avg_price = ?, fees = ? "
                "WHERE id = ? AND user_id = ?",
                (float(quantity), float(avg_price), float(fees),
                 int(position_id), uid),
            )
        conn.commit()
        return True
    except Exception:
        return False
    finally:
        conn.close()


# --- Portfolio settings (cash, devise…) — user-scoped ---

def get_portfolio_setting(key: str, default=None, user_id: Optional[str] = None):
    """Récupère une valeur de paramètre portefeuille pour l'utilisateur courant."""
    uid = _resolve_user(user_id)
    conn = get_connection()
    row = conn.execute(
        "SELECT value FROM portfolio_settings WHERE user_id = ? AND key = ?",
        (uid, key),
    ).fetchone()
    conn.close()
    if not row:
        return default
    return row[0]


def set_portfolio_setting(key: str, value, user_id: Optional[str] = None) -> None:
    """Enregistre ou met à jour un paramètre portefeuille pour l'utilisateur courant."""
    uid = _resolve_user(user_id)
    conn = get_connection()
    conn.execute(
        """INSERT INTO portfolio_settings (user_id, key, value, updated_at)
           VALUES (?, ?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(user_id, key) DO UPDATE SET
             value = excluded.value,
             updated_at = CURRENT_TIMESTAMP""",
        (uid, key, str(value) if value is not None else None),
    )
    conn.commit()
    conn.close()


def get_portfolio_cash(user_id: Optional[str] = None) -> float:
    v = get_portfolio_setting("cash", "0", user_id=user_id)
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def set_portfolio_cash(amount: float, user_id: Optional[str] = None) -> None:
    set_portfolio_setting("cash", float(amount), user_id=user_id)


# --- Investor Profile (user-scoped) ---

def save_investor_profile(profile: dict, user_id: Optional[str] = None):
    uid = _resolve_user(user_id)
    conn = get_connection()
    conn.execute(
        """INSERT INTO investor_profile (user_id, risk_profile, horizon, budget,
           preferred_sectors, preferred_tickers, excluded_tickers, objective, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(user_id) DO UPDATE SET
           risk_profile=excluded.risk_profile, horizon=excluded.horizon,
           budget=excluded.budget, preferred_sectors=excluded.preferred_sectors,
           preferred_tickers=excluded.preferred_tickers,
           excluded_tickers=excluded.excluded_tickers,
           objective=excluded.objective, updated_at=CURRENT_TIMESTAMP""",
        (
            uid,
            profile.get("risk_profile"),
            profile.get("horizon"),
            profile.get("budget"),
            json.dumps(profile.get("preferred_sectors", []), ensure_ascii=False),
            json.dumps(profile.get("preferred_tickers", []), ensure_ascii=False),
            json.dumps(profile.get("excluded_tickers", []), ensure_ascii=False),
            profile.get("objective"),
        ),
    )
    conn.commit()
    conn.close()


def get_investor_profile(user_id: Optional[str] = None) -> Optional[dict]:
    uid = _resolve_user(user_id)
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM investor_profile WHERE user_id = ?",
        (uid,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    profile = dict(row)
    for key in ("preferred_sectors", "preferred_tickers", "excluded_tickers"):
        if profile.get(key):
            try:
                profile[key] = json.loads(profile[key])
            except (json.JSONDecodeError, TypeError):
                profile[key] = []
    return profile


# --- Market Data (auto-scraped from sikafinance) ---

def save_market_data(data: dict):
    """Sauvegarde les donnees de marche scrapees pour un titre.

    IMPORTANT : on utilise COALESCE pour `company_name` et `sector` afin de
    NE PAS ecraser ces colonnes avec NULL quand le scraper ne les fournit pas
    (cas frequent : fetch_daily_quotes_brvm ne renvoie que prix/variation).
    Sinon chaque sync intraday remettait sector a NULL pour tous les tickers.
    """
    conn = get_connection()
    conn.execute(
        """INSERT INTO market_data (ticker, company_name, sector, price, variation,
           market_cap, beta, rsi, dps, dividend_history, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(ticker) DO UPDATE SET
           company_name=COALESCE(excluded.company_name, market_data.company_name),
           sector=COALESCE(excluded.sector, market_data.sector),
           price=excluded.price, variation=excluded.variation,
           market_cap=COALESCE(excluded.market_cap, market_data.market_cap),
           beta=COALESCE(excluded.beta, market_data.beta),
           rsi=COALESCE(excluded.rsi, market_data.rsi),
           dps=COALESCE(excluded.dps, market_data.dps),
           dividend_history=COALESCE(excluded.dividend_history, market_data.dividend_history),
           updated_at=CURRENT_TIMESTAMP""",
        (
            data.get("ticker"), data.get("name"), data.get("sector"),
            data.get("price"), data.get("variation"),
            data.get("market_cap"), data.get("beta"), data.get("rsi"),
            data.get("dps"),
            json.dumps(data.get("dividend_history", []), ensure_ascii=False),
        ),
    )
    conn.commit()
    conn.close()


def get_market_data(ticker: str = None) -> pd.DataFrame:
    """Recupere les donnees de marche. Si ticker=None, retourne tout."""
    conn = get_connection()
    if ticker:
        df = read_sql_df("SELECT * FROM market_data WHERE ticker=?", params=(ticker,)
        )
    else:
        df = read_sql_df("SELECT * FROM market_data ORDER BY ticker")
    conn.close()
    return df


@_maybe_cache_data(ttl=300)
def get_all_stocks_for_analysis() -> pd.DataFrame:
    """
    Fusionne market_data (auto-scrape) + fundamentals (Excel/manuel).
    Priorite aux donnees fondamentales quand disponibles.

    Résultat mis en cache 5 min (st.cache_data) pour éviter de re-requêter
    Supabase à chaque navigation de page.
    """
    best_year = _best_year_subquery()
    df = read_sql_df(f"""
        SELECT
            COALESCE(f.ticker, m.ticker) as ticker,
            COALESCE(f.company_name, m.company_name) as company_name,
            COALESCE(f.sector, m.sector) as sector,
            COALESCE(m.price, f.price) as price,
            m.variation, m.market_cap, m.beta, m.rsi,
            COALESCE(f.shares, m.shares) as shares,
            f.revenue, f.net_income, f.equity, f.total_debt,
            f.ebit, f.interest_expense, f.cfo, f.capex,
            f.dividends_total, f.total_assets, f.eps, f.per,
            COALESCE(f.dps, m.dps) as dps,
            m.dividend_yield as market_dividend_yield,
            f.fiscal_year, COALESCE(f.float_pct, m.float_pct) as float_pct,
            f.revenue_n3, f.revenue_n2, f.revenue_n1, f.revenue_n0,
            f.net_income_n3, f.net_income_n2, f.net_income_n1, f.net_income_n0,
            f.dps_n3, f.dps_n2, f.dps_n1, f.dps_n0,
            CASE WHEN f.ticker IS NOT NULL THEN 1 ELSE 0 END as has_fundamentals,
            m.updated_at as market_updated_at
        FROM market_data m
        LEFT JOIN (
            SELECT ff.* FROM fundamentals ff
            INNER JOIN ({best_year}) latest
            ON ff.ticker = latest.ticker AND ff.fiscal_year = latest.max_year
        ) f ON m.ticker = f.ticker
        UNION
        SELECT
            f.ticker, f.company_name, f.sector, f.price,
            NULL as variation, NULL as market_cap, NULL as beta, NULL as rsi,
            f.shares, f.revenue, f.net_income, f.equity, f.total_debt,
            f.ebit, f.interest_expense, f.cfo, f.capex,
            f.dividends_total, f.total_assets, f.eps, f.per, f.dps,
            NULL as market_dividend_yield,
            f.fiscal_year, f.float_pct,
            f.revenue_n3, f.revenue_n2, f.revenue_n1, f.revenue_n0,
            f.net_income_n3, f.net_income_n2, f.net_income_n1, f.net_income_n0,
            f.dps_n3, f.dps_n2, f.dps_n1, f.dps_n0,
            1 as has_fundamentals,
            NULL as market_updated_at
        FROM fundamentals f
        INNER JOIN ({best_year}) latest
        ON f.ticker = latest.ticker AND f.fiscal_year = latest.max_year
        WHERE f.ticker NOT IN (SELECT ticker FROM market_data)
        ORDER BY ticker
    """)

    # Fallback : pour les champs bilan/cashflow manquants, chercher la dernière
    # année antérieure qui les a. Recalcule aussi l'historique N-3..N0 à la volée
    # (mix des lignes DB) pour que le changement d'année de référence n'efface
    # pas les graphes historiques.
    BS_FIELDS = ("equity", "total_debt", "total_assets", "cfo", "capex",
                 "ebit", "interest_expense", "dividends_total")
    HIST_SUFFIXES = ["n0", "n1", "n2", "n3"]
    # Ne garder que les colonnes qui existent réellement (évite UndefinedColumn
    # sur Postgres si une colonne a été ajoutée après la migration).
    # Conn fraîche pour éviter une conn idle fermée par le pooler Supabase.
    _conn_cols = get_connection()
    try:
        _existing_cols = set(_table_columns(_conn_cols, "fundamentals"))
    finally:
        try:
            _conn_cols.close()
        except Exception:
            pass
    _bs_available = tuple(f for f in BS_FIELDS if f in _existing_cols)

    if not df.empty:
        # ── Optim : éviter N+1 sur Postgres/Supabase. On charge TOUT l'historique
        # fundamentals en une seule requête, puis on fait le fallback en mémoire.
        _select_cols = ", ".join(["ticker", "fiscal_year", "revenue", "net_income", "dps"]
                                  + list(_bs_available))
        try:
            full_fund = read_sql_df(
                f"SELECT {_select_cols} FROM fundamentals ORDER BY ticker, fiscal_year DESC"
            )
        except Exception:
            full_fund = pd.DataFrame()

        # Index par ticker → liste de rows triées DESC par année
        fund_by_ticker = {}
        if not full_fund.empty:
            for tkr, grp in full_fund.groupby("ticker"):
                fund_by_ticker[tkr] = grp.sort_values("fiscal_year", ascending=False)

        for i, row in df.iterrows():
            ticker = row["ticker"]
            ref_year = row.get("fiscal_year")
            if not ticker or not ref_year:
                continue
            sub = fund_by_ticker.get(ticker)
            if sub is None or sub.empty:
                continue

            # Fallback bilan/cashflow : 1ère année ≤ ref_year avec valeur non-nulle
            sub_ref = sub[sub["fiscal_year"] <= ref_year]
            missing = [f for f in _bs_available if pd.isna(row.get(f)) or row.get(f) is None]
            for fld in missing:
                if fld not in sub_ref.columns:
                    continue
                non_null = sub_ref[sub_ref[fld].notna() & (sub_ref[fld] != 0)]
                if not non_null.empty:
                    df.at[i, fld] = non_null.iloc[0][fld]

            # Historique N-3..N0 : 4 années les plus récentes avec revenue/net_income
            hist = sub_ref[
                (sub_ref["revenue"].notna() & (sub_ref["revenue"] != 0))
                | (sub_ref["net_income"].notna() & (sub_ref["net_income"] != 0))
            ].head(4)
            for idx, (_, hr) in enumerate(hist.iterrows()):
                suf = HIST_SUFFIXES[idx]
                if pd.notna(hr.get("revenue")) and hr.get("revenue"):
                    df.at[i, f"revenue_{suf}"] = hr["revenue"]
                if pd.notna(hr.get("net_income")) and hr.get("net_income"):
                    df.at[i, f"net_income_{suf}"] = hr["net_income"]
                if pd.notna(hr.get("dps")) and hr.get("dps"):
                    df.at[i, f"dps_{suf}"] = hr["dps"]

    return df


# --- Publications Tracking ---

def save_publication(pub: dict) -> bool:
    """Sauvegarde une publication. Retourne True si c'est une nouvelle."""
    conn = get_connection()
    try:
        cur = conn.execute(
            """INSERT INTO publications (ticker, title, pub_type, period, url, pub_date)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(ticker, title) DO NOTHING""",
            (pub.get("ticker"), pub.get("title"), pub.get("type"),
             pub.get("period"), pub.get("url"), pub.get("date")),
        )
        # cursor.rowcount : compatible SQLite + Postgres (évite SELECT changes())
        is_new = bool(getattr(cur, "rowcount", 0) and cur.rowcount > 0)
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        is_new = False
    conn.close()
    return is_new


def get_publications(ticker: str = None, only_new: bool = False) -> pd.DataFrame:
    conn = get_connection()
    query = "SELECT * FROM publications"
    params = []
    conditions = []
    if ticker:
        conditions.append("ticker=?")
        params.append(ticker)
    if only_new:
        conditions.append("is_new=1")
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_at DESC"
    df = read_sql_df(query, params=params)
    conn.close()
    return df


def mark_publications_read(ticker: str = None):
    conn = get_connection()
    if ticker:
        conn.execute("UPDATE publications SET is_new=0 WHERE ticker=?", (ticker,))
    else:
        conn.execute("UPDATE publications SET is_new=0")
    conn.commit()
    conn.close()


def get_pending_publications(recent_days: Optional[int] = 7) -> pd.DataFrame:
    """Retourne les publications qui ne sont PAS encore intégrées en base.

    Args:
        recent_days: si fourni, ne retourne que les publications des N derniers jours
                     (pub_date ou created_at). Par défaut 7 jours. Passer None pour tout.

    Critères de "pending":
    - Rapports annuels dont la fiscal_year est > max(fundamentals.fiscal_year) du ticker
    - Publications trimestrielles/semestrielles non présentes dans quarterly_data
    - Toute publication marquée is_new=1 (nouvellement scrapée)
    """
    from datetime import datetime, timedelta

    conn = get_connection()

    date_filter_sql = ""
    params = {}
    if recent_days is not None and recent_days > 0:
        cutoff = (datetime.now() - timedelta(days=recent_days)).strftime("%Y-%m-%d")
        date_filter_sql = (
            " AND (p.pub_date >= :cutoff OR "
            "      (p.pub_date IS NULL AND date(p.created_at) >= :cutoff)) "
        )
        params["cutoff"] = cutoff

    try:
        df = read_sql_df(f"""
            WITH fund_max AS (
                SELECT ticker, MAX(fiscal_year) AS latest_year
                FROM fundamentals
                WHERE revenue IS NOT NULL AND revenue != 0
                GROUP BY ticker
            ),
            quart_present AS (
                SELECT DISTINCT ticker, fiscal_year FROM quarterly_data
            )
            SELECT p.id, p.ticker, p.title, p.pub_type, p.period,
                   p.fiscal_year, p.url, p.pub_date, p.is_new,
                   fm.latest_year AS current_year_in_db,
                   CASE
                      WHEN p.is_new = 1 THEN 'nouveau'
                      WHEN p.pub_type = 'annuel'
                           AND p.fiscal_year IS NOT NULL
                           AND (fm.latest_year IS NULL OR p.fiscal_year > fm.latest_year)
                           THEN 'annuel_non_integre'
                      WHEN p.pub_type IN ('trimestriel', 'semestriel')
                           AND qp.ticker IS NULL
                           THEN 'trimestriel_a_verifier'
                      ELSE NULL
                   END AS pending_reason
            FROM publications p
            LEFT JOIN fund_max fm ON p.ticker = fm.ticker
            LEFT JOIN quart_present qp
                   ON qp.ticker = p.ticker
                  AND qp.fiscal_year = p.fiscal_year
            WHERE COALESCE(p.ignored, 0) = 0
              AND (
                p.is_new = 1
                OR (p.pub_type = 'annuel'
                    AND p.fiscal_year IS NOT NULL
                    AND (fm.latest_year IS NULL OR p.fiscal_year > fm.latest_year))
                OR (p.pub_type IN ('trimestriel', 'semestriel')
                    AND qp.ticker IS NULL)
              )
            {date_filter_sql}
            ORDER BY p.pub_date DESC, p.created_at DESC
            """, params=params,)
    except Exception:
        df = pd.DataFrame()
    conn.close()
    return df


def get_data_gaps() -> pd.DataFrame:
    """Détecte les données probablement manquantes par rapport au cycle de publication
    BRVM/UEMOA. Renvoie un DataFrame avec :
    - ticker, name, sector
    - latest_year_in_db : dernière année fiscale trouvée en base
    - expected_latest : année attendue actuellement (N-1 si on est après le 30 avril, sinon N-2)
    - missing_annual : bool — rapport annuel manquant
    - missing_quarter : str — trimestre le plus récent attendu mais absent
    - gap_type : 'annuel' / 'trimestriel' / 'semestriel'
    """
    from datetime import datetime
    now = datetime.now()
    year = now.year
    month = now.month

    # Deadline annual = 30 avril (publication des comptes N-1 au plus tard fin avril).
    # Si nous sommes après le 30 avril de l'année N → année N-1 doit être disponible.
    expected_annual = year - 1 if month >= 5 else year - 2

    # Trimestres attendus — règle UEMOA : 45 jours après fin de trimestre.
    # Q1 (fin mars) → attendu vers mi-mai → on l'attend à partir de juin.
    # Q2/S1 (fin juin) → attendu vers mi-août → on l'attend à partir de septembre.
    # Q3 (fin sept) → attendu vers mi-nov → on l'attend à partir de décembre.
    # Q4/annuel (fin déc) → attendu vers avril → géré par expected_annual.
    expected_quarter_year = year
    if month >= 12:
        expected_quarter = 3
    elif month >= 9:
        expected_quarter = 2
    elif month >= 6:
        expected_quarter = 1
    else:
        expected_quarter = 4
        expected_quarter_year = year - 1

    conn = get_connection()
    tickers_df = read_sql_df("""
        SELECT md.ticker, md.company_name, md.sector
        FROM market_data md
        WHERE md.price > 0
        ORDER BY md.ticker
    """)

    if tickers_df.empty:
        conn.close()
        return pd.DataFrame()

    # ── Optim : 4 requêtes agrégées (au lieu de 48 × 4 = 192 round-trips) ──
    # 1) Dernière année de fundamentals par ticker (toute row, même partielle).
    # On ne filtre PAS sur revenue : le parser PDF rate parfois ce champ
    # mais l'integration reste consideree faite si la row existe (les
    # autres champs comme net_income/eps/dps ont ete extraits).
    latest_fund = read_sql_df(
        """SELECT ticker, MAX(fiscal_year) AS latest
           FROM fundamentals
           GROUP BY ticker"""
    )
    latest_map = dict(zip(latest_fund["ticker"], latest_fund["latest"])) if not latest_fund.empty else {}

    # 2) Dernier trimestre en DB par ticker + set des quarters utilisés
    # (utile pour détecter la cadence : ticker qui n'a jamais eu Q=1 ou Q=3
    # publie probablement seulement semestriellement)
    qtr = read_sql_df(
        """SELECT ticker, MAX(fiscal_year * 10 + quarter) AS latest_q
           FROM quarterly_data
           GROUP BY ticker"""
    )
    qtr_map = {}
    if not qtr.empty:
        for _, qr in qtr.iterrows():
            code = qr["latest_q"] or 0
            qtr_map[qr["ticker"]] = (code // 10, code % 10)

    qtr_hist = read_sql_df(
        """SELECT ticker, quarter, COUNT(*) AS n
           FROM quarterly_data GROUP BY ticker, quarter"""
    )
    # Pour chaque ticker : set des quarters vus historiquement
    quarters_seen = {}
    if not qtr_hist.empty:
        for _, r in qtr_hist.iterrows():
            quarters_seen.setdefault(r["ticker"], set()).add(int(r["quarter"]))

    # 3) Publications annuelles : dernière année par ticker
    pub_annual = read_sql_df(
        """SELECT ticker, MAX(fiscal_year) AS latest_year
           FROM publications
           WHERE pub_type = 'annuel' AND fiscal_year IS NOT NULL
           GROUP BY ticker"""
    )
    pub_annual_map = dict(zip(pub_annual["ticker"], pub_annual["latest_year"])) if not pub_annual.empty else {}

    # 4) Publications trimestrielles récentes NON encore intégrées.
    # Filtre : exclut les tickers pour lesquels quarterly_data contient déjà
    # une ligne pour la même fiscal_year que la publication.
    pub_trim = read_sql_df(
        """SELECT p.ticker, p.title, p.fiscal_year
           FROM publications p
           INNER JOIN (
               SELECT ticker, MAX(pub_date) AS max_dt
               FROM publications
               WHERE pub_type IN ('trimestriel','semestriel')
               GROUP BY ticker
           ) mx ON p.ticker = mx.ticker AND p.pub_date = mx.max_dt
           WHERE p.pub_type IN ('trimestriel','semestriel')
             AND NOT EXISTS (
                SELECT 1 FROM quarterly_data qd
                WHERE qd.ticker = p.ticker
                  AND qd.fiscal_year = p.fiscal_year
             )"""
    )
    pub_trim_map = dict(zip(pub_trim["ticker"], pub_trim["title"])) if not pub_trim.empty else {}

    # 5) Gaps ignorés (lecture en masse)
    # NB : sur Postgres, fiscal_year NULL revient en NaN via pandas.
    # `is None` ne matche pas les NaN → on utilise pd.isna().
    ignored = read_sql_df("SELECT ticker, gap_type, fiscal_year FROM ignored_gaps")

    def _is_permanent(fy) -> bool:
        return fy is None or pd.isna(fy)

    def _as_int_year(fy):
        try:
            return int(fy) if not _is_permanent(fy) else None
        except (TypeError, ValueError):
            return None

    ignored_annual = {(r["ticker"], _as_int_year(r["fiscal_year"]))
                      for _, r in ignored.iterrows()
                      if r.get("gap_type") == "annuel"
                      and not _is_permanent(r["fiscal_year"])}
    ignored_annual_any = {r["ticker"] for _, r in ignored.iterrows()
                          if r.get("gap_type") == "annuel"
                          and _is_permanent(r["fiscal_year"])}
    ignored_trim = {r["ticker"] for _, r in ignored.iterrows()
                    if r.get("gap_type") == "trimestriel"}

    rows = []
    for _, tr in tickers_df.iterrows():
        ticker = tr["ticker"]
        latest = latest_map.get(ticker)

        missing_annual = latest is None or latest < expected_annual
        q_tuple = qtr_map.get(ticker)
        has_quarterly_history = q_tuple is not None
        has_expected_quarter = False
        if q_tuple:
            q_year, q_q = q_tuple
            if q_year > expected_quarter_year or (
                q_year == expected_quarter_year and q_q >= expected_quarter
            ):
                has_expected_quarter = True
        # Détection cadence : si le ticker n'a JAMAIS publié le quarter attendu
        # historiquement (ex : publie seulement Q=2/S1 et Q=4/S2, jamais Q=1/T1
        # ni Q=3/T3), on ne l'alerte pas sur ce quarter.
        seen = quarters_seen.get(ticker, set())
        publishes_this_quarter = (expected_quarter in seen) if seen else True
        missing_quarter = (
            f"{expected_quarter_year} Q{expected_quarter}"
            if has_quarterly_history and not has_expected_quarter
            and publishes_this_quarter
            else None
        )

        pub_gap = pub_annual_map.get(ticker)
        published_but_missing_year = None
        if pub_gap and (latest is None or pub_gap > latest):
            published_but_missing_year = int(pub_gap)
            missing_annual = True

        pub_q_title = pub_trim_map.get(ticker)

        # Filtre ignored
        if missing_annual and (
            ticker in ignored_annual_any or (ticker, expected_annual) in ignored_annual
        ):
            missing_annual = False
            published_but_missing_year = None
        if (missing_quarter or pub_q_title) and ticker in ignored_trim:
            missing_quarter = None
            pub_q_title = None

        if not (missing_annual or missing_quarter or published_but_missing_year or pub_q_title):
            continue

        gap_type = []
        if missing_annual:
            gap_type.append("annuel")
        if missing_quarter:
            gap_type.append("trimestriel")
        if pub_q_title:
            gap_type.append("publication trimestrielle récente")
            if not missing_quarter:
                missing_quarter = pub_q_title[:60]
        rows.append({
            "ticker": ticker,
            "name": tr["company_name"],
            "sector": tr["sector"],
            "latest_year_in_db": latest,
            "expected_latest": expected_annual,
            "missing_annual": missing_annual,
            "missing_quarter": missing_quarter,
            "gap_type": ", ".join(gap_type) if gap_type else "",
            "published_year": published_but_missing_year,
        })

    conn.close()
    return pd.DataFrame(rows)


# --- Actions sur les publications / écarts ---

def ignore_publication(pub_id: int) -> bool:
    """Marque une publication comme ignorée."""
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE publications SET ignored = 1, is_new = 0 WHERE id = ?",
            (pub_id,),
        )
        conn.commit()
        ok = True
    except sqlite3.Error:
        ok = False
    conn.close()
    return ok


def delete_publication(pub_id: int) -> bool:
    """Supprime définitivement une publication."""
    conn = get_connection()
    try:
        conn.execute("DELETE FROM publications WHERE id = ?", (pub_id,))
        conn.commit()
        ok = True
    except sqlite3.Error:
        ok = False
    conn.close()
    return ok


def mark_publication_integrated(pub_id: int) -> bool:
    """Marque comme traitée (is_new=0)."""
    conn = get_connection()
    try:
        conn.execute("UPDATE publications SET is_new = 0 WHERE id = ?", (pub_id,))
        conn.commit()
        ok = True
    except sqlite3.Error:
        ok = False
    conn.close()
    return ok


def ignore_gap(ticker: str, gap_type: str, fiscal_year: int = None,
                reason: str = None) -> bool:
    """Marque un écart calculé comme 'non applicable'."""
    conn = get_connection()
    try:
        conn.execute(
            """INSERT OR IGNORE INTO ignored_gaps (ticker, gap_type, fiscal_year, reason)
               VALUES (?, ?, ?, ?)""",
            (ticker, gap_type, fiscal_year, reason),
        )
        conn.commit()
        ok = True
    except sqlite3.Error:
        ok = False
    conn.close()
    return ok


def restore_gap(ticker: str, gap_type: str, fiscal_year: int = None) -> bool:
    """Annule le marquage 'non applicable'."""
    conn = get_connection()
    try:
        if fiscal_year is None:
            conn.execute(
                "DELETE FROM ignored_gaps WHERE ticker = ? AND gap_type = ?",
                (ticker, gap_type),
            )
        else:
            conn.execute(
                "DELETE FROM ignored_gaps WHERE ticker = ? AND gap_type = ? AND fiscal_year = ?",
                (ticker, gap_type, fiscal_year),
            )
        conn.commit()
        ok = True
    except sqlite3.Error:
        ok = False
    conn.close()
    return ok


def list_ignored_gaps() -> pd.DataFrame:
    conn = get_connection()
    df = read_sql_df("SELECT * FROM ignored_gaps ORDER BY created_at DESC")
    conn.close()
    return df


# --- Auto-clearance & dormant detection ---

def _mark_pubs_processed(conn, ticker: str, fiscal_year: int,
                         pub_types: tuple) -> int:
    """Marque comme traitées (is_new=0) les publications correspondant à
    (ticker, fiscal_year, pub_type in pub_types). Retourne le nombre marqué.
    Prend une connexion existante pour rester dans la transaction en cours."""
    if not ticker or not fiscal_year or not pub_types:
        return 0
    placeholders = ",".join(["?"] * len(pub_types))
    cur = conn.execute(
        f"""UPDATE publications SET is_new = 0
            WHERE is_new = 1 AND ticker = ?
              AND fiscal_year = ?
              AND pub_type IN ({placeholders})""",
        (ticker, fiscal_year) + tuple(pub_types),
    )
    return cur.rowcount or 0


def reconcile_publications(non_financial_cleanup_days: int = 60) -> dict:
    """Sweep : marque comme traitées les publications déjà intégrées, et
    archive les publications non-financières anciennes.

    - annuel/etats_financiers → is_new=0 si fundamentals(ticker, fiscal_year) existe
    - trimestriel/semestriel → is_new=0 si quarterly_data(ticker, fiscal_year) existe
    - gouvernance/corporate/autre/dividende avec pub_date > N jours → is_new=0

    Idempotent : peut tourner à chaque snapshot quotidien sans effet secondaire.
    Retourne un résumé des compteurs.
    """
    from datetime import datetime, timedelta
    conn = get_connection()
    result = {"annual_cleared": 0, "quarterly_cleared": 0,
              "non_financial_archived": 0}
    try:
        # Publications annuelles/EF intégrées → is_new=0
        cur = conn.execute("""
            UPDATE publications SET is_new = 0
            WHERE is_new = 1
              AND pub_type IN ('annuel', 'etats_financiers')
              AND fiscal_year IS NOT NULL
              AND ticker IN (
                  SELECT ticker FROM fundamentals
                  WHERE fundamentals.fiscal_year = publications.fiscal_year
                    AND revenue IS NOT NULL
              )
        """)
        result["annual_cleared"] = cur.rowcount or 0

        # Publications trimestrielles/semestrielles intégrées → is_new=0
        cur = conn.execute("""
            UPDATE publications SET is_new = 0
            WHERE is_new = 1
              AND pub_type IN ('trimestriel', 'semestriel')
              AND fiscal_year IS NOT NULL
              AND ticker IN (
                  SELECT ticker FROM quarterly_data
                  WHERE quarterly_data.fiscal_year = publications.fiscal_year
              )
        """)
        result["quarterly_cleared"] = cur.rowcount or 0

        # Publications non-financières anciennes → archivage automatique
        cutoff = (datetime.now() - timedelta(days=non_financial_cleanup_days)
                  ).strftime("%Y-%m-%d")
        cur = conn.execute("""
            UPDATE publications SET is_new = 0
            WHERE is_new = 1
              AND pub_type NOT IN ('annuel', 'etats_financiers',
                                    'trimestriel', 'semestriel')
              AND pub_date IS NOT NULL
              AND pub_date < ?
        """, (cutoff,))
        result["non_financial_archived"] = cur.rowcount or 0

        conn.commit()
    finally:
        conn.close()
    return result


def get_dormant_tickers(threshold_months: int = 12) -> pd.DataFrame:
    """Retourne les tickers candidats à un marquage inactif.

    Un ticker est considéré dormant si TOUTES ces conditions sont vraies :
    - Coté (market_data.price > 0)
    - Dernier signal de publication (max entre publications.pub_date et
      report_links proxy-date) date de plus de `threshold_months` mois
    - Pas de report_links pour l'exercice courant ni l'exercice précédent

    Exclut les tickers déjà marqués inactifs (ignored_gaps avec fiscal_year NULL).
    Colonnes : ticker, name, last_pub_date, months_since, total_pubs, reason.
    """
    from datetime import datetime, timedelta
    now = datetime.now()
    cutoff = (now - timedelta(days=int(threshold_months * 30.44))
              ).strftime("%Y-%m-%d")
    current_fy = now.year - 1 if now.month >= 5 else now.year - 2

    md = read_sql_df(
        "SELECT ticker, company_name AS name FROM market_data WHERE price > 0"
    )
    if md.empty:
        return md

    pub_agg = read_sql_df("""
        SELECT ticker,
               MAX(pub_date) AS last_pub_date,
               COUNT(*) AS total_pubs
        FROM publications
        GROUP BY ticker
    """)

    # Signal alternatif : dernière année fiscale ciblée par un report_link.
    # Approximation : on considère qu'un report FY=Y est publié en avril Y+1.
    rl_agg = read_sql_df("""
        SELECT ticker, MAX(fiscal_year) AS max_report_fy
        FROM report_links
        WHERE fiscal_year IS NOT NULL
        GROUP BY ticker
    """)

    df = md.merge(pub_agg, on="ticker", how="left").merge(
        rl_agg, on="ticker", how="left"
    )
    df["total_pubs"] = df["total_pubs"].fillna(0).astype(int)
    df["last_pub_date"] = pd.to_datetime(df["last_pub_date"], errors="coerce")

    def _report_proxy(fy):
        return pd.Timestamp(f"{int(fy) + 1}-04-30") if pd.notna(fy) else pd.NaT
    df["report_proxy_date"] = df["max_report_fy"].apply(_report_proxy)

    # Signal effectif = plus récent entre pub_date et report_proxy_date
    df["last_signal"] = df[["last_pub_date", "report_proxy_date"]].max(axis=1)

    cutoff_ts = pd.Timestamp(cutoff)
    no_signal = df["last_signal"].isna()
    stale = df["last_signal"].notna() & (df["last_signal"] < cutoff_ts)

    # Filtre supplémentaire : si report_links a un rapport pour l'exercice
    # courant ou N-1, on considère le titre actif quand même
    has_recent_reports = df["max_report_fy"].fillna(0) >= (current_fy - 1)
    df = df[(no_signal | stale) & ~has_recent_reports].copy()
    if df.empty:
        return df

    df["reason"] = df.apply(
        lambda r: "aucune_publication"
        if pd.isna(r["last_signal"])
        else "publications_arretees",
        axis=1,
    )
    df["months_since"] = df["last_signal"].apply(
        lambda d: round((now - d).days / 30.44, 1) if pd.notna(d) else None
    )

    permanent = read_sql_df("""
        SELECT DISTINCT ticker FROM ignored_gaps
        WHERE gap_type = 'annuel' AND fiscal_year IS NULL
    """)
    if not permanent.empty:
        df = df[~df["ticker"].isin(permanent["ticker"])]

    # Colonne de compat : last_pub_date affichée = last_signal
    df["last_pub_date"] = df["last_signal"]
    df = df[
        ["ticker", "name", "last_pub_date", "months_since",
         "total_pubs", "reason"]
    ]
    df = df.sort_values(
        ["reason", "last_pub_date"],
        ascending=[True, True],
        na_position="first",
    ).reset_index(drop=True)
    return df


def mark_ticker_inactive(ticker: str, reason: str = "Titre dormant") -> bool:
    """Marque un ticker comme inactif : insère dans ignored_gaps avec
    fiscal_year=NULL (ignore permanent pour toute future année).
    Utilisé pour les sociétés qui ne publient plus."""
    return ignore_gap(ticker, "annuel", fiscal_year=None, reason=reason)


# --- Extraction attempts (traçabilité des échecs PDF) ---

def get_recent_extraction_attempts(limit: int = 50,
                                     status_filter: Optional[str] = None) -> pd.DataFrame:
    """Retourne les dernières tentatives d'extraction PDF, avec par défaut
    les échecs uniquement. `status_filter` : 'success', 'no_pdf',
    'parser_empty', 'error', 'save_error' ou None (tout)."""
    where = ""
    params = []
    if status_filter:
        where = "WHERE status = ?"
        params = [status_filter]
    else:
        where = "WHERE status != 'success'"
    query = f"""
        SELECT a.id, a.publication_id, a.ticker, a.fiscal_year, a.pub_type,
               a.report_link_url, a.status, a.error_message,
               a.extracted_summary, a.attempted_at,
               p.title AS pub_title
        FROM extraction_attempts a
        LEFT JOIN publications p ON p.id = a.publication_id
        {where}
        ORDER BY a.attempted_at DESC
        LIMIT ?
    """
    params.append(limit)
    df = read_sql_df(query, params=tuple(params))
    return df


def get_latest_extraction_attempt_per_pub() -> pd.DataFrame:
    """Pour chaque publication_id, retourne la dernière tentative connue.
    Utile pour afficher un badge de statut par pub dans le fil d'actualité.
    Requête portable (fonctionne SQLite + Postgres)."""
    df = read_sql_df("""
        SELECT a.publication_id, a.status, a.error_message,
               a.extracted_summary, a.attempted_at
        FROM extraction_attempts a
        INNER JOIN (
            SELECT publication_id, MAX(attempted_at) AS max_ts
            FROM extraction_attempts
            WHERE publication_id IS NOT NULL
            GROUP BY publication_id
        ) latest
        ON latest.publication_id = a.publication_id
        AND latest.max_ts = a.attempted_at
    """)
    return df


# --- Dividendes encaissés ---

def save_dividend(data: dict) -> int:
    """Enregistre un dividende encaissé.

    data attend : ticker, payment_date, net_amount (obligatoires),
    gross_amount, withholding_tax, fiscal_year, notes (optionnels),
    user_id (défaut = current_user_id())."""
    conn = get_connection()
    uid = data.get("user_id") or current_user_id()
    cursor = conn.execute(
        """INSERT INTO dividends
           (user_id, ticker, payment_date, net_amount, gross_amount,
            withholding_tax, fiscal_year, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (uid, data.get("ticker"), data.get("payment_date"),
         data.get("net_amount"), data.get("gross_amount"),
         data.get("withholding_tax"), data.get("fiscal_year"),
         data.get("notes")),
    )
    conn.commit()
    row_id = cursor.lastrowid
    conn.close()
    return row_id


def get_dividends(user_id: Optional[str] = None) -> pd.DataFrame:
    """Retourne les dividendes de l'utilisateur, triés par date décroissante."""
    uid = user_id or current_user_id()
    df = read_sql_df(
        """SELECT id, ticker, payment_date, net_amount, gross_amount,
                  withholding_tax, fiscal_year, notes, created_at
           FROM dividends WHERE user_id = ?
           ORDER BY payment_date DESC, id DESC""",
        params=(uid,),
    )
    return df


def update_dividend(dividend_id: int, data: dict,
                     user_id: Optional[str] = None) -> bool:
    """Met à jour un dividende. Champs éditables : ticker, payment_date,
    net_amount, gross_amount, withholding_tax, fiscal_year, notes."""
    uid = user_id or current_user_id()
    editable = ["ticker", "payment_date", "net_amount", "gross_amount",
                "withholding_tax", "fiscal_year", "notes"]
    updates = [(c, data[c]) for c in editable if c in data]
    if not updates:
        return False
    set_clause = ", ".join([f"{c} = ?" for c, _ in updates])
    params = tuple(v for _, v in updates) + (dividend_id, uid)
    conn = get_connection()
    try:
        conn.execute(
            f"""UPDATE dividends SET {set_clause}, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND user_id = ?""",
            params,
        )
        conn.commit()
        ok = True
    except Exception:
        ok = False
    conn.close()
    return ok


def delete_dividend(dividend_id: int, user_id: Optional[str] = None) -> bool:
    """Supprime un dividende (uniquement si l'utilisateur en est propriétaire)."""
    uid = user_id or current_user_id()
    conn = get_connection()
    try:
        conn.execute(
            "DELETE FROM dividends WHERE id = ? AND user_id = ?",
            (dividend_id, uid),
        )
        conn.commit()
        ok = True
    except Exception:
        ok = False
    conn.close()
    return ok


def get_total_dividends_received(user_id: Optional[str] = None,
                                  ticker: Optional[str] = None) -> float:
    """Somme des dividendes nets encaissés (par ticker ou global)."""
    uid = user_id or current_user_id()
    if ticker:
        df = read_sql_df(
            "SELECT COALESCE(SUM(net_amount), 0) AS total FROM dividends "
            "WHERE user_id = ? AND ticker = ?",
            params=(uid, ticker),
        )
    else:
        df = read_sql_df(
            "SELECT COALESCE(SUM(net_amount), 0) AS total FROM dividends "
            "WHERE user_id = ?",
            params=(uid,),
        )
    return float(df["total"].iloc[0]) if not df.empty else 0.0


# --- Frais de compte (Type 2 : non liés à un titre spécifique) ---

ACCOUNT_FEE_CATEGORIES = [
    "droits_garde",
    "tenue_compte",
    "virement",
    "commission_dividende",
    "conversion",
    "autre",
]


def save_account_fee(data: dict, debit_cash: bool = True) -> int:
    """Enregistre un frais de compte (non lié à un titre).

    Si debit_cash=True (défaut), le cash portfolio est débité du montant.
    data attend : fee_date, category, amount (obligatoires),
    notes (optionnel), user_id (défaut = current_user_id())."""
    conn = get_connection()
    uid = data.get("user_id") or current_user_id()
    cursor = conn.execute(
        """INSERT INTO account_fees
           (user_id, fee_date, category, amount, notes)
           VALUES (?, ?, ?, ?, ?)""",
        (uid, data.get("fee_date"), data.get("category"),
         float(data.get("amount", 0)), data.get("notes")),
    )
    conn.commit()
    row_id = cursor.lastrowid
    conn.close()

    if debit_cash and data.get("amount"):
        try:
            current_cash = get_portfolio_cash(user_id=uid)
            set_portfolio_cash(current_cash - float(data["amount"]), user_id=uid)
        except Exception:
            pass  # non bloquant

    return row_id


def get_account_fees(user_id: Optional[str] = None) -> pd.DataFrame:
    """Retourne les frais de compte de l'utilisateur, triés par date décroissante."""
    uid = user_id or current_user_id()
    df = read_sql_df(
        """SELECT id, fee_date, category, amount, notes, created_at
           FROM account_fees WHERE user_id = ?
           ORDER BY fee_date DESC, id DESC""",
        params=(uid,),
    )
    return df


def delete_account_fee(fee_id: int, refund_cash: bool = True,
                        user_id: Optional[str] = None) -> bool:
    """Supprime un frais. Si refund_cash=True (défaut), recrédite le cash
    du montant."""
    uid = user_id or current_user_id()
    conn = get_connection()
    amount = 0
    try:
        row = conn.execute(
            "SELECT amount FROM account_fees WHERE id = ? AND user_id = ?",
            (fee_id, uid),
        ).fetchone()
        if row:
            amount = float(row[0] or 0)
        conn.execute(
            "DELETE FROM account_fees WHERE id = ? AND user_id = ?",
            (fee_id, uid),
        )
        conn.commit()
        ok = True
    except Exception:
        ok = False
    conn.close()

    if ok and refund_cash and amount:
        try:
            current_cash = get_portfolio_cash(user_id=uid)
            set_portfolio_cash(current_cash + amount, user_id=uid)
        except Exception:
            pass

    return ok


def get_total_account_fees(user_id: Optional[str] = None) -> float:
    """Somme des frais de compte payés (Type 2 uniquement)."""
    uid = user_id or current_user_id()
    df = read_sql_df(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM account_fees "
        "WHERE user_id = ?",
        params=(uid,),
    )
    return float(df["total"].iloc[0]) if not df.empty else 0.0


# --- Quarterly Data ---

def save_quarterly_data(data: dict) -> int:
    conn = get_connection()
    cursor = conn.execute(
        """INSERT INTO quarterly_data (ticker, fiscal_year, quarter, periode,
           revenue, net_income, ebit, source, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(ticker, fiscal_year, quarter) DO UPDATE SET
           periode=COALESCE(excluded.periode, quarterly_data.periode),
           revenue=excluded.revenue, net_income=excluded.net_income,
           ebit=excluded.ebit, source=excluded.source, notes=excluded.notes""",
        (data.get("ticker"), data.get("fiscal_year"), data.get("quarter"),
         data.get("periode"),
         data.get("revenue"), data.get("net_income"), data.get("ebit"),
         data.get("source"), data.get("notes")),
    )
    conn.commit()
    row_id = cursor.lastrowid

    # Auto-clearance : marque les publications trimestrielles/semestrielles
    # correspondantes comme traitées.
    ticker = data.get("ticker")
    fy = data.get("fiscal_year")
    if ticker and fy:
        try:
            _mark_pubs_processed(conn, ticker, fy,
                                  ("trimestriel", "semestriel"))
            conn.commit()
        except Exception:
            pass  # non bloquant

    conn.close()
    return row_id


def get_quarterly_data(ticker: str, fiscal_year: int = None) -> pd.DataFrame:
    conn = get_connection()
    if fiscal_year:
        df = read_sql_df("SELECT * FROM quarterly_data WHERE ticker=? AND fiscal_year=? ORDER BY quarter", params=(ticker, fiscal_year),
        )
    else:
        df = read_sql_df("SELECT * FROM quarterly_data WHERE ticker=? ORDER BY fiscal_year DESC, quarter", params=(ticker,),
        )
    conn.close()
    return df


def get_publication_calendar() -> pd.DataFrame:
    """Calendrier des publications attendues avec statut RÉEL.

    Croise 4 sources :
    - Liste des tickers cotés (market_data + fundamentals)
    - Publications reçues (`publications`)
    - Données intégrées (`fundamentals`, `quarterly_data`)
    - Gaps ignorés / titres dormants (`ignored_gaps`)

    Périodes considérées par ticker : Annuel N-1 (avril), T1 N (mai),
    S1 N (septembre), T3 N (novembre) — cycle standard BRVM/UEMOA.
    Ajoute aussi Annuel N-2 si toujours pas intégré (retard structurel).

    Statuts retournés :
    - `integre`        : publication reçue ET données en base → vert
    - `a_integrer`     : publication reçue mais données pas encore en base → orange
    - `en_retard`      : mois attendu passé ET aucune publication → rouge
    - `a_venir`        : mois attendu pas encore atteint → gris
    (les titres dormants et gaps ignorés sont filtrés en sortie.)
    """
    from datetime import datetime
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    # 1) Tickers cotés
    tickers = read_sql_df("""
        SELECT ticker,
               MAX(company_name) AS company_name,
               MAX(sector) AS sector
        FROM (
            SELECT ticker, company_name, sector FROM fundamentals
            UNION ALL
            SELECT ticker, company_name, sector FROM market_data
        ) u
        GROUP BY ticker
        ORDER BY ticker
    """)
    if tickers.empty:
        return pd.DataFrame()

    from config import load_tickers
    config_tickers = {t["ticker"]: t for t in load_tickers()}

    # 2) Publications reçues par (ticker, pub_type, fiscal_year)
    pubs = read_sql_df("""
        SELECT ticker, pub_type, fiscal_year, MAX(pub_date) AS pub_date,
               MAX(id) AS pub_id, MAX(url) AS url, MAX(title) AS title
        FROM publications
        WHERE COALESCE(ignored, 0) = 0
          AND pub_type IN ('annuel', 'etats_financiers', 'trimestriel', 'semestriel')
        GROUP BY ticker, pub_type, fiscal_year
    """)
    pubs_by_key = {}
    if not pubs.empty:
        for _, r in pubs.iterrows():
            fy = r.get("fiscal_year")
            try:
                fy_int = int(fy) if fy is not None and not pd.isna(fy) else None
            except (TypeError, ValueError):
                fy_int = None
            key = (r["ticker"], (r["pub_type"] or "").lower(), fy_int)
            pubs_by_key[key] = {
                "pub_id": r.get("pub_id"),
                "pub_date": r.get("pub_date"),
                "url": r.get("url"),
                "title": r.get("title"),
            }

    # 3) Intégration annuelle : max(fiscal_year) avec AU MOINS un chiffre clé
    # non-null (revenue, net_income ou total_assets). Assouplissement pour les
    # rapports où l'un des 3 n'a pas pu être extrait mais les autres oui.
    fund = read_sql_df("""
        SELECT ticker, MAX(fiscal_year) AS max_fy
        FROM fundamentals
        WHERE revenue IS NOT NULL
           OR net_income IS NOT NULL
           OR total_assets IS NOT NULL
        GROUP BY ticker
    """)
    fund_max = dict(zip(fund["ticker"], fund["max_fy"])) if not fund.empty else {}

    # 4) Intégration trimestrielle/semestrielle : set(ticker, fy)
    quart = read_sql_df("SELECT DISTINCT ticker, fiscal_year FROM quarterly_data")
    quart_set: set = set()
    if not quart.empty:
        for _, r in quart.iterrows():
            try:
                quart_set.add((r["ticker"], int(r["fiscal_year"])))
            except (TypeError, ValueError):
                continue

    # 5) Ignore rules
    ignored = read_sql_df("SELECT ticker, gap_type, fiscal_year FROM ignored_gaps")
    ignored_permanent = set()   # ticker → ignore all annual (fiscal_year NULL)
    ignored_annual_year = set() # (ticker, year) → ignore just this year
    ignored_trim_permanent = set()
    if not ignored.empty:
        for _, r in ignored.iterrows():
            gt = r.get("gap_type")
            fy = r.get("fiscal_year")
            is_null = fy is None or pd.isna(fy)
            if gt == "annuel":
                if is_null:
                    ignored_permanent.add(r["ticker"])
                else:
                    try:
                        ignored_annual_year.add((r["ticker"], int(fy)))
                    except (TypeError, ValueError):
                        pass
            elif gt == "trimestriel":
                if is_null:
                    ignored_trim_permanent.add(r["ticker"])

    def _match_pub(ticker: str, pub_types: tuple, fy: int):
        """Cherche une publication qui matche parmi les types acceptés."""
        for pt in pub_types:
            key = (ticker, pt, fy)
            if key in pubs_by_key:
                return pubs_by_key[key]
        return None

    def _compute_status(published: bool, integrated: bool,
                         expected_month: int, year_diff: int):
        """year_diff = 0 si année courante, <0 si passé (retard structurel).

        Règle : si la data est intégrée, c'est « integre » quel que soit
        l'état de publications (les data peuvent venir d'Excel, saisie
        manuelle, ou d'un scraper qui n'a pas capturé l'annonce mais a
        eu le PDF via un autre canal). L'important pour le user, c'est
        d'avoir la data en base."""
        if integrated:
            return "integre"
        if published:
            return "a_integrer"
        # Ni publié ni intégré : selon date attendue
        if year_diff < 0:
            return "en_retard"  # année passée non publiée = très en retard
        if expected_month < current_month:
            return "en_retard"
        if expected_month == current_month:
            return "attendu_ce_mois"
        return "a_venir"

    calendar = []
    for _, row in tickers.iterrows():
        ticker = row["ticker"]
        if ticker in ignored_permanent:
            continue  # ticker inactif, on ne l'affiche pas
        cfg = config_tickers.get(ticker, {})
        name = row["company_name"] or cfg.get("name", ticker)
        sector = row["sector"] or cfg.get("sector", "")

        # Périodes candidates : Annuel N-2 (si en retard structurel), Annuel N-1,
        # puis T1/S1/T3 de l'année courante
        periods = []
        # Annuel N-2 seulement si pas encore intégré (sinon inutile de le montrer)
        fy_n2 = current_year - 2
        if (fund_max.get(ticker) or 0) < fy_n2 and (ticker, fy_n2) not in ignored_annual_year:
            periods.append(dict(
                period=f"Annuel {fy_n2}", type="annuel",
                fy=fy_n2, expected_month=4, year_diff=-2,
            ))
        # Annuel N-1
        fy_n1 = current_year - 1
        if (ticker, fy_n1) not in ignored_annual_year:
            periods.append(dict(
                period=f"Annuel {fy_n1}", type="annuel",
                fy=fy_n1, expected_month=4, year_diff=-1,
            ))
        # Trimestriels/semestriels de l'année courante (skip si ticker ne publie pas)
        if ticker not in ignored_trim_permanent:
            for period_label, pt, fy_p, month, ydiff in [
                (f"T1 {current_year}", "trimestriel", current_year, 5, 0),
                (f"S1 {current_year}", "semestriel", current_year, 9, 0),
                (f"T3 {current_year}", "trimestriel", current_year, 11, 0),
            ]:
                periods.append(dict(
                    period=period_label, type=pt, fy=fy_p,
                    expected_month=month, year_diff=ydiff,
                ))

        for p in periods:
            # Chercher publication (annuel accepte 'annuel' ou 'etats_financiers')
            if p["type"] == "annuel":
                pub_types = ("annuel", "etats_financiers")
                integrated = (fund_max.get(ticker) or 0) >= p["fy"]
            else:
                pub_types = (p["type"],)
                integrated = (ticker, p["fy"]) in quart_set
            pub = _match_pub(ticker, pub_types, p["fy"])
            published = pub is not None

            status = _compute_status(published, integrated,
                                       p["expected_month"], p["year_diff"])

            # Ne pas afficher les "à venir" pour ne pas encombrer le calendrier
            # (on montre déjà de quoi être occupé). Seul « attendu ce mois »
            # est utile côté action.
            if status == "a_venir":
                continue

            calendar.append({
                "ticker": ticker,
                "company_name": name,
                "sector": sector,
                "period": p["period"],
                "type": p["type"],
                "fiscal_year": p["fy"],
                "expected_month": p["expected_month"],
                "status": status,
                "pub_id": pub.get("pub_id") if pub else None,
                "pub_date": pub.get("pub_date") if pub else None,
                "pub_url": pub.get("url") if pub else None,
            })

    return pd.DataFrame(calendar)


# --- Qualitative Notes (user-scoped) ---

def save_qualitative_note(ticker: str, category: str, content: str,
                          source: str = None, note_date: str = None,
                          user_id: Optional[str] = None) -> int:
    uid = _resolve_user(user_id)
    conn = get_connection()
    cursor = conn.execute(
        """INSERT INTO qualitative_notes
           (user_id, ticker, category, content, source, note_date)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (uid, ticker, category, content, source, note_date),
    )
    conn.commit()
    row_id = cursor.lastrowid
    conn.close()
    return row_id


def get_qualitative_notes(ticker: str, user_id: Optional[str] = None) -> pd.DataFrame:
    uid = _resolve_user(user_id)
    conn = get_connection()
    df = read_sql_df("""SELECT * FROM qualitative_notes
           WHERE ticker = ? AND user_id = ?
           ORDER BY created_at DESC""", params=(ticker, uid),
    )
    conn.close()
    return df


def delete_qualitative_note(note_id: int, user_id: Optional[str] = None):
    uid = _resolve_user(user_id)
    conn = get_connection()
    conn.execute(
        "DELETE FROM qualitative_notes WHERE id = ? AND user_id = ?",
        (note_id, uid),
    )
    conn.commit()
    conn.close()


# --- Report Links ---

def save_report_link(data: dict) -> bool:
    conn = get_connection()
    try:
        conn.execute(
            """INSERT INTO report_links (ticker, title, report_type, fiscal_year, url, source)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(ticker, url) DO UPDATE SET
               title=excluded.title, report_type=excluded.report_type,
               fiscal_year=excluded.fiscal_year""",
            (data.get("ticker"), data.get("title"), data.get("report_type"),
             data.get("fiscal_year"), data.get("url"), data.get("source", "brvm.org")),
        )
        conn.commit()
        saved = True
    except Exception:
        saved = False
    conn.close()
    return saved


def get_report_links(ticker: str = None) -> pd.DataFrame:
    conn = get_connection()
    if ticker:
        df = read_sql_df("SELECT * FROM report_links WHERE ticker=? ORDER BY fiscal_year DESC", params=(ticker,),
        )
    else:
        df = read_sql_df("SELECT * FROM report_links ORDER BY fiscal_year DESC, ticker")
    conn.close()
    return df


def seed_known_report_links():
    """Pre-charge les liens connus vers les rapports annuels 2023-2024 des 48 titres BRVM."""
    known_reports = [
        # === SONATEL ===
        {"ticker": "SNTS.sn", "title": "Etats financiers 2024 - Sonatel", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250221_-_etats_financiers_-_exercice_2024_-_sonatel_sn.pdf"},
        {"ticker": "SNTS.sn", "title": "Rapport annuel 2024 - Sonatel", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://sonatel.sn/wp-content/uploads/2025/04/SONATEL_RA-2024_A4-Digital.pdf"},
        {"ticker": "SNTS.sn", "title": "Rapport activites T3 2024 - Sonatel", "report_type": "rapport_trimestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20241025_-_rapport_dactivites_-_3eme_trimestre_2024_-_sonatel_sn.pdf"},
        {"ticker": "SNTS.sn", "title": "Etats financiers 2023 - Sonatel", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240222_-_etats_financiers_-_exercice_2023_-_sonatel_sn.pdf"},
        # === ORANGE CI ===
        {"ticker": "ORAC.ci", "title": "Rapport annuel integre 2024 - Orange CI", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250416_-_rapport_annuel_integre_-_exercice_2024_-_orange_ci.pdf"},
        {"ticker": "ORAC.ci", "title": "Etats financiers 2024 - Orange CI", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250221_-_etats_financiers_-_exercice_2024_-_orange_ci.pdf"},
        # === ECOBANK CI ===
        {"ticker": "ECOC.ci", "title": "Rapport activites 2024 - Ecobank CI", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/fr/ecobank-cote-divoire-rapport-dactivites-annuel-exercice-2024"},
        {"ticker": "ECOC.ci", "title": "Note de recherche 2024 - Ecobank CI", "report_type": "analyse", "fiscal_year": 2024, "url": "https://www.bridge-securities.com/images/app/contenu/312/VFNotedeRecherche2024ECOBANKCIFR.pdf"},
        # === NSIA BANQUE ===
        {"ticker": "NSBC.ci", "title": "Etats financiers certifies 2024 - NSIA Banque", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/en/nsia-banque-cote-divoire-financial-statements-certified-statutory-auditors-2024"},
        {"ticker": "NSBC.ci", "title": "Note de recherche - NSIA Banque CI", "report_type": "analyse", "fiscal_year": 2024, "url": "https://www.bridge-securities.com/images/app/contenu/332/NSIABANQUECINotedeRechercheFR.pdf"},
        {"ticker": "NSBC.ci", "title": "Etats financiers 2023 - NSIA Banque", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240328_-_etats_financiers_-_exercice_2023_-_nsia_banque_ci.pdf"},
        # === SGBCI ===
        {"ticker": "SGBC.ci", "title": "Rapport activites 2024 - SGBCI", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250207_-_rapport_dactivites_annuel_-_exercice_2024_-_societe_generale_ci.pdf"},
        {"ticker": "SGBC.ci", "title": "Page rapports BRVM - SGBCI", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/fr/rapports-societe-cotes/sgb-ci"},
        # === ETI TOGO ===
        {"ticker": "ETIT.tg", "title": "Etats financiers certifies 2024 - ETI TG", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250328_-_etats_financiers_certifies_par_les_commissaires_aux_comptes_-_exercice_2024_-_eti_tg.pdf"},
        {"ticker": "ETIT.tg", "title": "Etats financiers 2023 - ETI TG", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240502_-_etats_financiers_-_exercice_2023_-_eti_tg.pdf"},
        # === BICICI ===
        {"ticker": "BICC.ci", "title": "Etats financiers 2024 - BICICI", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250328_-_etats_financiers_-_exercice_2024_-_bici_ci.pdf"},
        {"ticker": "BICC.ci", "title": "Rapport activites S1 2024 - BICICI", "report_type": "rapport_semestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20241031_-_rapport_dactivites_-_1er_semestre_2024_-_bici_ci.pdf"},
        # === SIB CI ===
        {"ticker": "SIBC.ci", "title": "Rapport activites 2024 - SIB CI", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250305_-_rapport_dactivites_annuel_-_exercice_2024_-_sib_ci.pdf"},
        # === BOA SENEGAL ===
        {"ticker": "BOAS.sn", "title": "Etats financiers certifies 2024 - BOA Senegal", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250319_-_etats_financiers_certifies_par_les_cacs_-_exercice_2024_-_boa_senegal.pdf"},
        # === BOA NIGER ===
        {"ticker": "BOAN.ne", "title": "Etats financiers 2023 - BOA Niger", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240325_-_etats_financiers_-_exercice_2023_-_boa_niger.pdf"},
        # === PALM CI ===
        {"ticker": "PALC.ci", "title": "Etats financiers provisoires 2024 - Palm CI", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250417_-_etats_financiers_provisoires_valides_par_le_ca_-_exercice_2024_-_palm_ci.pdf"},
        # === FILTISAC ===
        {"ticker": "FTSC.ci", "title": "Etats financiers 2024 - Filtisac CI", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250523_-_etats_financiers_-_exercice_2024_-_filtisac_ci.pdf"},
        {"ticker": "FTSC.ci", "title": "Rapport activites T1 2024 - Filtisac", "report_type": "rapport_trimestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20240513_-_rapport_dactivites_-_1er_trimestre_2024_-_filtisac_ci_0.pdf"},
        # === SODECI ===
        {"ticker": "SDCC.ci", "title": "Etats financiers 2024 - SODECI", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250430_-_etats_financiers_-_exercice_2024_-_sode_ci.pdf"},
        {"ticker": "SDCC.ci", "title": "Rapport activites S1 2024 - SODECI", "report_type": "rapport_semestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20241031_-_rapport_dactivites_-_1er_semestre_2024_-_sode_ci.pdf"},
        # === CIE CI ===
        {"ticker": "CIEC.ci", "title": "Rapport activites S1 2024 certifie - CIE CI", "report_type": "rapport_semestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20241030_-_rapport_dactivites_certifie_par_les_cacs_-_1er_semestre_2024_-_cie_ci.pdf"},
        # === SERVAIR ABIDJAN ===
        {"ticker": "ABJC.ci", "title": "Etats financiers IFRS 2024 - Servair Abidjan", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250430_-_etats_financiers_-_norme_ifrs_-_exercice_2024_-_servair_abidjan_ci.pdf"},
        {"ticker": "ABJC.ci", "title": "Etats financiers 2023 - Servair Abidjan", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240819_-_etats_financiers_approuves_-_exercice_2023_-_servair_abidjan_ci.pdf"},
        # === ONATEL BF ===
        {"ticker": "ONTBF.bf", "title": "Rapport gestion 2024 - Onatel BF", "report_type": "rapport_annuel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250718_-_rapport_annuel_de_gestion_-_exercice_2024_-_onatel_bf.pdf"},
        {"ticker": "ONTBF.bf", "title": "Resultats financiers 2023 - Onatel BF", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240404_-_resultats_financiers_-_exercice_2023_-_onatel_bf.pdf"},
        # === ORAGROUP TOGO ===
        {"ticker": "ORGT.tg", "title": "Rapport activites S1 2024 - Oragroup", "report_type": "rapport_semestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20240930_-_rapport_dactivites_-_1er_semestre_2024_-_oragroup_tg.pdf"},
        # === SITAB ===
        {"ticker": "STBC.ci", "title": "Rapport activites S1 2024 - SITAB", "report_type": "rapport_semestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20241031_-_rapport_dactivites_-_1er_semestre_2024_-_sitab_ci.pdf"},
        # === SAFCA ===
        {"ticker": "SAFC.ci", "title": "Rapport activites T4 2024 - SAFCA", "report_type": "rapport_trimestriel", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20240624_-_rapport_dactivites_-_4eme_trimestre_2024_-_safca_ci.pdf"},
        # === LOTERIE NATIONALE BENIN ===
        {"ticker": "LNBB.bj", "title": "Etats financiers 2024 - LNB Benin", "report_type": "etats_financiers", "fiscal_year": 2024, "url": "https://www.brvm.org/sites/default/files/20250602_-_etats_financiers_-_exercice_2024_-_lnb_bn.pdf"},
        # === TOTAL CI ===
        {"ticker": "TTLC.ci", "title": "Etats financiers definitifs 2023 - TotalEnergies CI", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240910_-_etats_financiers_definitifs_2023_-_totalenergies_marketing_ci.pdf"},
        # === SUCRIVOIRE ===
        {"ticker": "SCRC.ci", "title": "Etats financiers 2023 - Sucrivoire", "report_type": "etats_financiers", "fiscal_year": 2023, "url": "https://www.brvm.org/sites/default/files/20240510_-_etats_financiers_-_exercice_2023_-_sucrivoire_ci.pdf"},
    ]
    for report in known_reports:
        save_report_link(report)


# --- Company Profiles & News ---

def save_company_profile(data: dict):
    """Sauvegarde ou met a jour le profil qualitatif d'une societe."""
    conn = get_connection()
    cols = ["ticker", "description", "business", "president", "dg", "dga",
            "phone", "fax", "address", "website",
            "major_shareholder", "major_shareholder_pct"]
    values = [data.get(c) for c in cols]
    placeholders = ", ".join(["?"] * len(cols))
    col_names = ", ".join(cols)
    update_clause = ", ".join(f"{c}=excluded.{c}" for c in cols if c != "ticker")

    conn.execute(
        f"""INSERT INTO company_profiles ({col_names})
            VALUES ({placeholders})
            ON CONFLICT(ticker) DO UPDATE SET {update_clause}, updated_at=CURRENT_TIMESTAMP""",
        values,
    )
    conn.commit()
    conn.close()


def get_company_profile(ticker: str) -> dict:
    """Recupere le profil qualitatif d'un titre."""
    conn = get_connection()
    row = conn.execute("SELECT * FROM company_profiles WHERE ticker = ?", (ticker,)).fetchone()
    conn.close()
    if row:
        return dict(row)
    return {}


def get_all_company_profiles() -> pd.DataFrame:
    """Recupere tous les profils."""
    conn = get_connection()
    try:
        df = read_sql_df("SELECT * FROM company_profiles ORDER BY ticker")
    except Exception:
        df = pd.DataFrame()
    conn.close()
    return df


def save_company_news(ticker: str, articles: list):
    """Sauvegarde les actualites d'un titre."""
    conn = get_connection()
    for art in articles:
        try:
            conn.execute(
                """INSERT OR IGNORE INTO company_news (ticker, title, url, article_date)
                   VALUES (?, ?, ?, ?)""",
                (ticker, art.get("title", ""), art.get("url", ""), art.get("date")),
            )
        except Exception:
            pass
    conn.commit()
    conn.close()


def get_company_news(ticker: str = None, limit: int = 50) -> pd.DataFrame:
    """Recupere les actualites. Si ticker=None, toutes les actus."""
    conn = get_connection()
    if ticker:
        df = read_sql_df("SELECT * FROM company_news WHERE ticker = ? ORDER BY created_at DESC LIMIT ?", params=(ticker, limit),
        )
    else:
        df = read_sql_df("SELECT * FROM company_news ORDER BY created_at DESC LIMIT ?", params=(limit,),
        )
    conn.close()
    return df


# --- Signal History (pour calibrage long terme) ---
# Modèle "événement": une ligne unique par événement.
# Règle de fusion: si le même (ticker, entry_type, signal_name, verdict) réapparaît
# dans les 7 jours suivant sa dernière occurrence, on met à jour la ligne existante.
# Sinon, on crée un nouvel événement (nouvelle ligne).

SIGNAL_EVENT_GAP_DAYS = 7


def _upsert_signal_event(
    conn: sqlite3.Connection,
    ticker: str,
    entry_type: str,
    signal_name: Optional[str],
    verdict: Optional[str],
    today: str,
    payload: dict,
) -> str:
    """Merge ou insère un événement selon la règle de gap.
    Retourne: 'new', 'updated' ou 'skip'."""
    from datetime import datetime, timedelta

    # Find the most recent row for this (ticker, entry_type, signal_name, verdict) key
    row = conn.execute(
        """SELECT id, last_seen_date, occurrence_count
           FROM signal_history
           WHERE ticker = ? AND entry_type = ?
             AND COALESCE(signal_name, '') = COALESCE(?, '')
             AND COALESCE(verdict, '') = COALESCE(?, '')
           ORDER BY last_seen_date DESC LIMIT 1""",
        (ticker, entry_type, signal_name, verdict),
    ).fetchone()

    today_dt = datetime.strptime(today, "%Y-%m-%d")

    if row:
        last_seen = row[1]
        try:
            last_dt = datetime.strptime(last_seen, "%Y-%m-%d")
        except (ValueError, TypeError):
            last_dt = None

        if last_dt is not None and (today_dt - last_dt).days <= SIGNAL_EVENT_GAP_DAYS:
            # Within gap - merge: update last_seen_date + increment count (idempotent on same day)
            if last_seen == today:
                return "skip"
            conn.execute(
                """UPDATE signal_history
                   SET last_seen_date = ?,
                       occurrence_count = occurrence_count + 1,
                       updated_at = CURRENT_TIMESTAMP
                   WHERE id = ?""",
                (today, row[0]),
            )
            return "updated"

    # New event
    conn.execute(
        """INSERT INTO signal_history
           (ticker, company_name, sector, entry_type,
            signal_type, signal_name, verdict,
            first_seen_date, last_seen_date, occurrence_count,
            price_at_start, signal_strength, signal_details,
            hybrid_score, fundamental_score, technical_score,
            stars, trend)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            ticker, payload.get("company_name"), payload.get("sector"), entry_type,
            payload.get("signal_type"), signal_name, verdict,
            today, today,
            payload.get("price_at_start"),
            payload.get("signal_strength"),
            payload.get("signal_details"),
            payload.get("hybrid_score"),
            payload.get("fundamental_score"),
            payload.get("technical_score"),
            payload.get("stars"),
            payload.get("trend"),
        ),
    )
    return "new"


def save_signal_snapshots(
    ticker: str,
    signals: list,
    price: Optional[float] = None,
    company_name: Optional[str] = None,
    sector: Optional[str] = None,
    snapshot_date: Optional[str] = None,
) -> int:
    """Enregistre les signaux d'un ticker pour aujourd'hui.
    Un signal qui réapparaît dans les 7 jours ne crée PAS de nouvelle ligne :
    il met à jour last_seen_date et incrémente occurrence_count.
    Si plus de 7 jours sans apparition → nouvel événement (nouvelle ligne).
    Retourne le nombre de NOUVEAUX événements créés."""
    from datetime import datetime
    if not signals:
        return 0
    snapshot_date = snapshot_date or datetime.now().strftime("%Y-%m-%d")
    conn = get_connection()
    new_events = 0
    for sig in signals:
        signal_name = sig.get("signal") or sig.get("name")
        payload = {
            "company_name": company_name, "sector": sector,
            "signal_type": sig.get("type"),
            "signal_strength": sig.get("strength"),
            "signal_details": sig.get("details"),
            "price_at_start": price,
        }
        try:
            status = _upsert_signal_event(
                conn, ticker, "signal", signal_name, None, snapshot_date, payload,
            )
            if status == "new":
                new_events += 1
        except sqlite3.Error:
            continue
    conn.commit()
    conn.close()
    return new_events


def save_recommendation_snapshot(
    ticker: str,
    recommendation: dict,
    hybrid_score: float,
    fundamental_score: float,
    technical_score: float,
    price: Optional[float] = None,
    trend: Optional[str] = None,
    company_name: Optional[str] = None,
    sector: Optional[str] = None,
    snapshot_date: Optional[str] = None,
) -> bool:
    """Enregistre un snapshot de recommandation hybride.
    Même règle de fusion que save_signal_snapshots :
    si le verdict est le même et apparu il y a ≤ 7 jours → merge.
    Sinon → nouvel événement.
    Retourne True si une nouvelle ligne a été créée."""
    from datetime import datetime
    snapshot_date = snapshot_date or datetime.now().strftime("%Y-%m-%d")
    verdict = recommendation.get("verdict") if isinstance(recommendation, dict) else None
    stars = recommendation.get("stars") if isinstance(recommendation, dict) else None
    payload = {
        "company_name": company_name, "sector": sector,
        "price_at_start": price, "trend": trend,
        "hybrid_score": hybrid_score,
        "fundamental_score": fundamental_score,
        "technical_score": technical_score,
        "stars": stars,
    }
    conn = get_connection()
    try:
        status = _upsert_signal_event(
            conn, ticker, "recommendation",
            "hybrid_recommendation", verdict, snapshot_date, payload,
        )
        conn.commit()
        is_new = status == "new"
    except sqlite3.Error:
        is_new = False
    conn.close()
    return is_new


def delete_signal_event(event_id: int) -> bool:
    """Supprime un événement de l'historique par son ID. Retourne True si OK."""
    conn = get_connection()
    try:
        cur = conn.execute("DELETE FROM signal_history WHERE id = ?", (event_id,))
        conn.commit()
        deleted = cur.rowcount > 0
    except sqlite3.Error:
        deleted = False
    conn.close()
    return deleted


def delete_signal_events_bulk(event_ids: list) -> int:
    """Supprime plusieurs événements. Retourne le nombre supprimé."""
    if not event_ids:
        return 0
    conn = get_connection()
    placeholders = ",".join(["?"] * len(event_ids))
    try:
        cur = conn.execute(
            f"DELETE FROM signal_history WHERE id IN ({placeholders})",
            event_ids,
        )
        conn.commit()
        n = cur.rowcount
    except sqlite3.Error:
        n = 0
    conn.close()
    return n


def get_signal_history(
    ticker: Optional[str] = None,
    entry_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> pd.DataFrame:
    """Récupère les événements signaux/recommandations.
    Filtres sur last_seen_date (pour voir 'encore actifs dernièrement')."""
    conn = get_connection()
    conditions = []
    params = []
    if ticker:
        conditions.append("ticker = ?")
        params.append(ticker)
    if entry_type:
        conditions.append("entry_type = ?")
        params.append(entry_type)
    if start_date:
        conditions.append("last_seen_date >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("first_seen_date <= ?")
        params.append(end_date)

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    query = f"""SELECT * FROM signal_history {where_clause}
                ORDER BY last_seen_date DESC, ticker"""
    df = read_sql_df(query, params=params)
    conn.close()
    return df


def compute_signal_performance(df: pd.DataFrame) -> pd.DataFrame:
    """Calcule la performance depuis first_seen_date à 1M/3M/6M/1A.
    Ajoute aussi perf_since_start (du prix de départ jusqu'au prix actuel)."""
    if df.empty:
        return df
    from datetime import datetime, timedelta

    df = df.copy()
    df["perf_1m"] = None
    df["perf_3m"] = None
    df["perf_6m"] = None
    df["perf_1a"] = None
    df["perf_since_start"] = None
    df["current_price"] = None

    conn = get_connection()
    try:
        for idx, row in df.iterrows():
            ticker = row["ticker"]
            start_date = row.get("first_seen_date")
            ref_price = row.get("price_at_start")

            if not start_date or not ticker:
                continue

            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            except (ValueError, TypeError):
                continue

            # Fallback to price_cache if no stored start price
            if not ref_price:
                r = conn.execute(
                    """SELECT close FROM price_cache
                       WHERE ticker = ? AND date <= ?
                       ORDER BY date DESC LIMIT 1""",
                    (ticker, start_date),
                ).fetchone()
                if r:
                    ref_price = r[0]

            if not ref_price:
                continue

            # Latest price
            latest = conn.execute(
                "SELECT close FROM price_cache WHERE ticker = ? ORDER BY date DESC LIMIT 1",
                (ticker,),
            ).fetchone()
            if latest and latest[0]:
                df.at[idx, "current_price"] = latest[0]
                df.at[idx, "perf_since_start"] = (latest[0] - ref_price) / ref_price

            # Horizons (measured from first_seen_date)
            horizons = [("perf_1m", 30), ("perf_3m", 91), ("perf_6m", 182), ("perf_1a", 365)]
            for col, days in horizons:
                target = start_dt + timedelta(days=days)
                r = conn.execute(
                    """SELECT close FROM price_cache
                       WHERE ticker = ? AND date >= ?
                       ORDER BY date ASC LIMIT 1""",
                    (ticker, target.strftime("%Y-%m-%d")),
                ).fetchone()
                if r and r[0] and ref_price:
                    df.at[idx, col] = (r[0] - ref_price) / ref_price
    finally:
        conn.close()

    return df


# Initialize DB on import
init_db()


# Libelles lisibles des parcs clients telecoms
PARCS_LIBELLES = {
    "parc_mobile": "Clients mobile",
    "parc_mobile_money": "Clients mobile money",
    "parc_fibre": "Clients fibre",
    "parc_fixe_broadband": "Clients fixe haut débit",
    "parc_data_mobile": "Clients data mobile",
    "parc_4g": "Clients 4G actifs",
    "parc_total": "Parc total",
}


def save_telecom_parcs(ticker: str, fiscal_year, periode, parcs: dict,
                       source_url: str = None) -> int:
    """Enregistre les parcs clients d'un operateur pour une periode.

    Le parc est le moteur du chiffre d'affaires : le suivre separement permet
    de voir un recul du mobile masque par la croissance de la fibre.
    """
    if not parcs:
        return 0
    conn = get_connection()
    ecrits = 0
    try:
        for parc, detail in parcs.items():
            valeur = detail.get("valeur") if isinstance(detail, dict) else detail
            variation = detail.get("variation") if isinstance(detail, dict) else None
            if not valeur:
                continue
            conn.execute(
                """INSERT INTO telecom_parcs
                   (ticker, fiscal_year, periode, parc, valeur, variation, source_url)
                   VALUES (?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT (ticker, fiscal_year, periode, parc) DO UPDATE SET
                     valeur = excluded.valeur, variation = excluded.variation,
                     source_url = excluded.source_url,
                     updated_at = CURRENT_TIMESTAMP""",
                (ticker, fiscal_year, periode or "", parc, valeur, variation,
                 source_url),
            )
            ecrits += 1
        conn.commit()
    except Exception:
        conn.rollback()
    finally:
        conn.close()
    return ecrits


def get_telecom_parcs(ticker: str):
    """Derniers parcs connus d'un operateur, du plus recent au plus ancien."""
    try:
        return read_sql_df(
            """SELECT fiscal_year, periode, parc, valeur, variation
               FROM telecom_parcs WHERE ticker = ?
               ORDER BY fiscal_year DESC, periode DESC""",
            params=(ticker,),
        )
    except Exception:
        return None
